#!/usr/bin/env python3
"""
VIP 数据导入脚本
从 C:\eBoss\Local\VIP档案维护06001.xlsx 导入会员数据到 samsung_ops.db

功能：
1. 读取 Excel VIP 数据
2. 门店名称模糊匹配到 stores 表
3. 基于消费金额自动计算会员等级
4. 批量插入 members 表（去重 by phone）
5. 生成模拟购买记录到 member_purchases 表
6. 建立标签体系并分配会员标签
"""

import sqlite3
import openpyxl
import re
import random
from datetime import datetime, timedelta
from pathlib import Path

# ── 配置 ──────────────────────────────────────────────
EXCEL_PATH = r"C:\eBoss\Local\VIP档案维护06001.xlsx"
DB_PATH = r"D:\Workbuudy\samsung-ops\data\samsung_ops.db"

# 会员等级规则（基于累计消费金额）
LEVEL_RULES = [
    (50000, "黑卡"),
    (20000, "钻卡"),
    (5000,  "金卡"),
    (1000,  "银卡"),
    (0,     "普通"),
]

def calc_level(total_spent):
    for threshold, level in LEVEL_RULES:
        if total_spent >= threshold:
            return level
    return "普通"

# 门店名称模糊匹配映射
STORE_ALIAS = {
    "华润万象汇": 2,
    "万象汇": 2,
    "兴义梦乐城": 3,
    "梦乐城": 3,
    "龙湾万达": 7,
    "万达三星专卖": 7,
    "遵义吾悦": 4,
    "吾悦三星": 4,
    "云南昭通": 8,
    "昭通": 8,
    "六盘水": 6,
    "万象城": 1,
    "曲靖万达": 5,
    "清镇吾悦": 9,
    "安顺": 10,
    "遵义夏兴": 11,
    "售后仓": None,  # 过滤掉售后仓记录
}

def match_store(name: str) -> int | None:
    if not name:
        return None
    name = str(name).strip()
    for alias, store_id in STORE_ALIAS.items():
        if alias in name:
            return store_id
    return None  # 未匹配到则设为 NULL

# ── 主逻辑 ────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  VIP 数据导入脚本")
    print("=" * 60)

    # 1. 加载 Excel
    print(f"\n[1/6] 读取 Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1
    print(f"       共 {total_rows} 条记录")

    # 2. 连接数据库
    print(f"\n[2/6] 连接数据库: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    cur = conn.cursor()

    # 获取已有机型列表（用于生成模拟购买记录）
    cur.execute("SELECT DISTINCT model_code FROM inventory WHERE model_code IS NOT NULL AND model_code != ''")
    available_models = [r[0] for r in cur.fetchall()]
    if not available_models:
        available_models = ["S25U", "ZF6", "S24U", "A56", "ZF5", "S25"]
    print(f"       可用机型: {len(available_models)} 种")

    # 3. 读取并转换数据
    print(f"\n[3/6] 解析 Excel 数据...")
    headers = [cell.value for cell in ws[1]]

    # 列索引
    idx = {
        "id":        headers.index("ID"),
        "vip_card":  headers.index("VIP卡号"),
        "name":      headers.index("顾客姓名"),
        "discount":  headers.index("折扣(%)"),
        "times":     headers.index("消费次数"),
        "amount":    headers.index("消费金额"),
        "points":    headers.index("消费积分"),
        "vip_type":  headers.index("VIP类型"),
        "store":     headers.index("店仓"),
        "gender":    headers.index("性别"),
        "id_card":   headers.index("身份证号"),
        "first_date": headers.index("首单"),
        "phone":     headers.index("手机"),
        "birthday":  headers.index("生日"),
        "join_date": headers.index("加入日期"),
        "expire":    headers.index("有效期到"),
        "recent":    headers.index("最近消费"),
        "tag":       headers.index("标签"),
        "deleted":   headers.index("作废"),
    }

    members = []
    skipped = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        # 跳过已作废（作废列值为"是"/True 时才跳过）
        deleted_val = row[idx["deleted"]]
        if deleted_val and str(deleted_val).strip() in ("是", "1", "true", "True", "TRUE", "√", "yes"):
            skipped += 1
            continue

        phone = str(row[idx["phone"]]).strip() if row[idx["phone"]] else ""
        if not phone or phone == "None":
            skipped += 1
            continue
        # 去掉空格和横线
        phone = re.sub(r'[\s\-]', '', phone)
        if len(phone) < 7 or len(phone) > 15:
            skipped += 1
            continue

        # 过滤售后仓
        store_name = row[idx["store"]]
        store_id = match_store(store_name)
        if store_id is None:
            skipped += 1
            continue

        total_spent = float(row[idx["amount"]] or 0)
        level = calc_level(total_spent)
        points = int(row[idx["points"]] or 0)
        consume_times = int(row[idx["times"]] or 0)

        # 解析日期
        def parse_date(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d")
            s = str(v).strip()
            for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                except Exception:
                    pass
            return None

        join_date = parse_date(row[idx["join_date"]])
        recent_date = parse_date(row[idx["recent"]])
        birthday = parse_date(row[idx["birthday"]])
        first_date = parse_date(row[idx["first_date"]])

        gender_map = {"男": "男", "女": "女"}
        gender = gender_map.get(str(row[idx["gender"]]).strip(), None)

        members.append({
            "vip_card_no":   str(row[idx["vip_card"]] or "")[:50],
            "name":          str(row[idx["name"]] or "未知")[:50],
            "phone":         phone[:20],
            "store_id":      store_id,
            "gender":        gender,
            "birthday":      birthday,
            "join_date":     join_date,
            "first_date":    first_date,
            "recent_date":   recent_date,
            "total_spent":   total_spent,
            "points":        points,
            "consume_times": consume_times,
            "level":         level,
            "tag_raw":       str(row[idx["tag"]] or "").strip(),
        })

        if len(members) % 2000 == 0:
            print(f"       已解析 {len(members)} 条...")

    wb.close()
    print(f"       解析完成: {len(members)} 条有效, {skipped} 条跳过")

    # 4. 去重（按 phone）
    print(f"\n[4/6] 去重处理...")
    seen = set()
    unique_members = []
    for m in members:
        if m["phone"] not in seen:
            seen.add(m["phone"])
            unique_members.append(m)
    print(f"       去重后: {len(unique_members)} 条（去掉 {len(members) - len(unique_members)} 条重复手机号）")

    # 5. 插入 members 表
    print(f"\n[5/6] 插入 members 表...")
    cur.execute("DELETE FROM members WHERE id > 9")  # 清空测试数据（保留前9条示例）
    conn.commit()

    insert_sql = """
        INSERT INTO members (name, phone, store_id, level, total_spent, points,
                            join_date, is_vip, is_active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, 1, 1, datetime('now'))
    """
    member_id_map = {}  # phone -> db id
    for m in unique_members:
        cur.execute(insert_sql, (
            m["name"], m["phone"], m["store_id"], m["level"],
            m["total_spent"], m["points"],
            m["join_date"] or datetime.now().strftime("%Y-%m-%d"),
        ))
        member_id_map[m["phone"]] = cur.lastrowid

    conn.commit()
    print(f"       插入 {len(unique_members)} 条会员记录")

    # 5b. 统计等级分布
    cur.execute("SELECT level, COUNT(*) FROM members GROUP BY level ORDER BY COUNT(*) DESC")
    print("       会员等级分布:")
    for row in cur.fetchall():
        print(f"         {row[0]}: {row[1]} 人")

    # 6. 生成模拟购买记录
    print(f"\n[6/6] 生成模拟购买记录...")
    cur.execute("DELETE FROM member_purchases")  # 清空旧数据
    conn.commit()

    purchase_sql = """
        INSERT INTO member_purchases
            (member_id, store_id, amount, points_earned, product_info,
             purchase_date, model_code, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
    """

    random.seed(42)
    total_purchases = 0
    now = datetime.now()

    for m in unique_members:
        mid = member_id_map.get(m["phone"])
        if not mid:
            continue

        times = max(1, m["consume_times"])  # 至少 1 次
        total = m["total_spent"]
        store_id = m["store_id"]
        join_dt = parse_date_str(m["join_date"]) or (now - timedelta(days=365*2))
        recent_dt = parse_date_str(m["recent_date"]) or now

        if recent_dt < join_dt:
            recent_dt = now

        # 生成每次购买的金额（总和 ≈ total）
        if times == 1:
            amounts = [total]
        else:
            # 随机分配，保证总和接近 total
            ratios = [random.uniform(0.3, 1.0) for _ in range(times)]
            s = sum(ratios)
            amounts = [total * r / s for r in ratios]

        # 生成购买日期（均匀分布在 join_date ~ recent_date）
        if times == 1:
            dates = [join_dt + (recent_dt - join_dt) * 0.7]
        else:
            step = (recent_dt - join_dt).days / (times + 1)
            dates = [join_dt + timedelta(days=int(step * (i + 1))) for i in range(times)]
            # 确保最后一个日期不超过 recent_dt
            dates[-1] = min(dates[-1], recent_dt)

        for i in range(times):
            amount = round(amounts[i], 2)
            pts = int(amount / 10)  # 10元 = 1积分
            model = random.choice(available_models)
            spec = random.choice(["8+256GB", "12+256GB", "12+512GB", "16+512GB"])
            color = random.choice(["星际黑", "冰川蓝", "羽夜银", "薄荷绿"])
            product_info = f"{model} {spec} {color}"
            date_str = dates[i].strftime("%Y-%m-%d")

            cur.execute(purchase_sql, (mid, store_id, amount, pts, product_info, date_str, model))
            total_purchases += 1

        if total_purchases % 5000 == 0:
            print(f"       已生成 {total_purchases} 条购买记录...")

    conn.commit()
    print(f"       生成 {total_purchases} 条模拟购买记录")

    # 7. 建立标签体系
    print(f"\n[Bonus] 建立标签体系...")
    cur.execute("DELETE FROM member_tag_defs")

    tag_defs = [
        ("高价值客户",  "red",    "累计消费 ≥ 20000元"),
        ("中价值客户",  "orange", "累计消费 5000~20000元"),
        ("潜力客户",    "blue",   "累计消费 1000~5000元"),
        ("新会员",      "green",  "加入 ≤ 90天"),
        ("沉睡客户",    "gray",   "超过180天未消费"),
        ("高频消费",    "purple", "消费次数 ≥ 10次"),
        ("男宾",        "cyan",   "性别男"),
        ("女宾",        "pink",   "性别女"),
    ]
    tag_id_map = {}
    for name, color, desc in tag_defs:
        cur.execute(
            "INSERT INTO member_tag_defs (name, color, description, created_at) VALUES (?, ?, ?, datetime('now'))",
            (name, color, desc)
        )
        tag_id_map[name] = cur.lastrowid

    # 分配标签
    print(f"       分配会员标签...")
    cur.execute("DELETE FROM member_tags")
    now_str = now.strftime("%Y-%m-%d")
    new_members_cutoff = (now - timedelta(days=90)).strftime("%Y-%m-%d")
    sleep_cutoff = (now - timedelta(days=180)).strftime("%Y-%m-%d")

    tag_assign_count = 0
    for m in unique_members:
        mid = member_id_map.get(m["phone"])
        if not mid:
            continue
        total = m["total_spent"]
        times = m["consume_times"]
        gender = m["gender"]
        join_date = m["join_date"] or ""
        recent_date = m["recent_date"] or ""

        tags_to_assign = []

        if total >= 20000:
            tags_to_assign.append("高价值客户")
        elif total >= 5000:
            tags_to_assign.append("中价值客户")
        elif total >= 1000:
            tags_to_assign.append("潜力客户")

        if join_date >= new_members_cutoff:
            tags_to_assign.append("新会员")

        if recent_date and recent_date < sleep_cutoff:
            tags_to_assign.append("沉睡客户")

        if times >= 10:
            tags_to_assign.append("高频消费")

        if gender == "男":
            tags_to_assign.append("男宾")
        elif gender == "女":
            tags_to_assign.append("女宾")

        for tag_name in tags_to_assign:
            tid = tag_id_map.get(tag_name)
            if tid:
                cur.execute(
                    "INSERT INTO member_tags (member_id, tag_id, created_at) VALUES (?, ?, datetime('now'))",
                    (mid, tid)
                )
                tag_assign_count += 1

    conn.commit()
    print(f"       分配 {tag_assign_count} 个标签关联")

    # 最终统计
    print(f"\n" + "=" * 60)
    print(f"  导入完成！")
    print(f"=" * 60)
    cur.execute("SELECT COUNT(*) FROM members")
    print(f"  members 表: {cur.fetchone()[0]} 条")
    cur.execute("SELECT COUNT(*) FROM member_purchases")
    print(f"  member_purchases 表: {cur.fetchone()[0]} 条")
    cur.execute("SELECT COUNT(*) FROM member_tag_defs")
    print(f"  member_tag_defs 表: {cur.fetchone()[0]} 个标签定义")
    cur.execute("SELECT COUNT(*) FROM member_tags")
    print(f"  member_tags 表: {cur.fetchone()[0]} 个标签关联")

    conn.close()
    print(f"\n✅ 全部完成！\n")

def parse_date_str(s: str) -> datetime | None:
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None

if __name__ == "__main__":
    main()
