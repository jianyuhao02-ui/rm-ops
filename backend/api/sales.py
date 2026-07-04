"""销售数据 API - eBoss数据驱动的月度任务报表（含对比分析、排名、导出）"""
import io
import os
import csv
from fastapi import APIRouter, Depends, HTTPException
from backend.dependencies import get_current_user, require_admin, require_manager_or_admin
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from backend.models.database import get_db

router = APIRouter()


class DailySalesRecord(BaseModel):
    store_id: int
    date: str  # YYYY-MM-DD
    phone_sales: float = 0
    ncme_sales: float = 0
    phone_qty: int = 0
    key_model_qty: int = 0
    accessory_sales: float = 0
    trade_in_qty: int = 0


class MonthlyTarget(BaseModel):
    store_id: int
    year: int
    month: int
    grade: str = ""
    phone_sales_target: float = 0
    ncme_target: float = 0
    phone_qty_target: int = 0
    key_model_target: int = 0
    accessory_target: float = 0
    trade_in_rate: float = 0


@router.get("/monthly")
async def get_monthly_sales(year: int, month: int,
    user: dict = Depends(get_current_user)):
    """获取指定月份各门店的销售进度（月度任务报表）"""

    db = await get_db()
    # 门店过滤条件
    if user["role"] == "staff" and user.get("store_id"):
        store_filter = "AND s.id=?"
        store_params = (user["store_id"],)
    else:
        store_filter = ""
        store_params = ()

    month_start = f"{year}-{month:02d}-01"
    next_month = f"{year}-{month + 1:02d}-01" if month < 12 else f"{year + 1}-01-01"

    # 一次查询：门店 LEFT JOIN 目标 LEFT JOIN 销售
    sql = f"""
        SELECT
            s.id as store_id, s.name as store_name, s.province,
            mt.grade,
            mt.phone_sales_target, mt.ncme_target, mt.phone_qty_target,
            mt.key_model_target, mt.accessory_target, mt.trade_in_rate,
            COALESCE(ds.total_phone_sales, 0) as phone_done,
            COALESCE(ds.total_ncme_sales, 0) as ncme_done,
            COALESCE(ds.total_phone_qty, 0) as phone_qty_done,
            COALESCE(ds.total_key_model_qty, 0) as km_done,
            COALESCE(ds.total_accessory_sales, 0) as acc_done,
            COALESCE(ds.total_trade_in_qty, 0) as trade_in_done
        FROM stores s
        LEFT JOIN monthly_targets mt ON mt.store_id = s.id AND mt.year=? AND mt.month=?
        LEFT JOIN (
            SELECT store_id,
                SUM(phone_sales) as total_phone_sales,
                SUM(ncme_sales) as total_ncme_sales,
                SUM(phone_qty) as total_phone_qty,
                SUM(key_model_qty) as total_key_model_qty,
                SUM(accessory_sales) as total_accessory_sales,
                SUM(trade_in_qty) as total_trade_in_qty
            FROM daily_sales
            WHERE date >= ? AND date < ?
            GROUP BY store_id
        ) ds ON ds.store_id = s.id
        WHERE s.is_active=1 {store_filter}
        ORDER BY s.sort_order
    """
    params = (year, month, month_start, next_month) + store_params
    cursor = await db.execute(sql, params)
    rows = await cursor.fetchall()

    # 来源统计（单次查询）
    cursor = await db.execute("""
        SELECT source, COUNT(*) as cnt, COUNT(DISTINCT date) as days
        FROM daily_sales WHERE date >= ? AND date < ? GROUP BY source
    """, (month_start, next_month))
    source_stats = {r["source"]: {"records": r["cnt"], "days": r["days"]} for r in await cursor.fetchall()}

    result = []
    for r in rows:
        phone_target = r["phone_sales_target"] or 0
        ncme_target = r["ncme_target"] or 0
        km_target = r["key_model_target"] or 0
        acc_target = r["accessory_target"] or 0
        phone_done = r["phone_done"]
        ncme_done = r["ncme_done"]
        km_done = r["km_done"]
        acc_done = r["acc_done"]
        phone_qty_done = r["phone_qty_done"]
        trade_in_done = r["trade_in_done"]

        result.append({
            "store_id": r["store_id"],
            "store_name": r["store_name"],
            "province": r["province"],
            "grade": r["grade"] or "",
            "phone_sales_target": phone_target,
            "phone_sales_done": phone_done,
            "phone_sales_rate": round(phone_done / phone_target * 100, 1) if phone_target > 0 else 0,
            "ncme_target": ncme_target,
            "ncme_done": ncme_done,
            "ncme_rate": round(ncme_done / ncme_target * 100, 1) if ncme_target > 0 else 0,
            "phone_qty_done": phone_qty_done,
            "key_model_target": km_target,
            "key_model_done": km_done,
            "key_model_rate": round(km_done / km_target * 100, 1) if km_target > 0 else 0,
            "accessory_target": acc_target,
            "accessory_done": acc_done,
            "accessory_rate": round(acc_done / acc_target * 100, 1) if acc_target > 0 else 0,
            "trade_in_done": trade_in_done,
            "trade_in_rate": round(trade_in_done / phone_qty_done * 100, 1) if phone_qty_done > 0 else 0,
        })

    return {
        "year": year, "month": month,
        "source_stats": source_stats,
        "stores": result
    }


@router.get("/daily-trend")
async def get_daily_trend(year: int, month: int,
    user: dict = Depends(get_current_user), store_id: int = None):
    """获取每日销售趋势（可按门店筛选）"""

    db = await get_db()
    month_start = f"{year}-{month:02d}-01"
    next_month = f"{year}-{month + 1:02d}-01" if month < 12 else f"{year + 1}-01-01"
    if store_id:
        rows = await db.execute("""
            SELECT date, SUM(phone_sales) as total_sales, SUM(phone_qty) as total_qty,
                   SUM(ncme_sales) as total_ncme, SUM(accessory_sales) as total_acc
            FROM daily_sales
            WHERE date >= ? AND date < ? AND store_id=?
            GROUP BY date ORDER BY date
        """, (month_start, next_month, store_id))
    else:
        rows = await db.execute("""
            SELECT date, SUM(phone_sales) as total_sales, SUM(phone_qty) as total_qty,
                   SUM(ncme_sales) as total_ncme, SUM(accessory_sales) as total_acc
            FROM daily_sales
            WHERE date >= ? AND date < ?
            GROUP BY date ORDER BY date
        """, (month_start, next_month))
    data = await rows.fetchall()
    return [{"date": r["date"], "sales": r["total_sales"] or 0,
             "qty": r["total_qty"] or 0, "ncme": r["total_ncme"] or 0,
             "acc": r["total_acc"] or 0} for r in data]


@router.get("/stores")
async def get_stores(
    user: dict = Depends(get_current_user)
):
    """获取门店列表"""

    db = await get_db()
    cursor = await db.execute("SELECT id, name, province FROM stores WHERE is_active=1 ORDER BY sort_order")
    return [dict(row) for row in await cursor.fetchall()]


@router.post("/record")
async def record_daily_sales(records: List[DailySalesRecord],
    user: dict = Depends(get_current_user)):
    """录入/更新每日销售数据（admin/manager）"""
    if user["role"] not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="权限不足")

    db = await get_db()
    upserted = 0
    for rec in records:
        # 店长只能录入本店数据
        if user["role"] == "manager" and user.get("store_id") and rec.store_id != user["store_id"]:
            continue

        await db.execute("""
            INSERT INTO daily_sales (date, store_id, phone_sales, ncme_sales, phone_qty,
                                      key_model_qty, accessory_sales, trade_in_qty)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(date, store_id)
            DO UPDATE SET phone_sales=excluded.phone_sales, ncme_sales=excluded.ncme_sales,
                         phone_qty=excluded.phone_qty, key_model_qty=excluded.key_model_qty,
                         accessory_sales=excluded.accessory_sales, trade_in_qty=excluded.trade_in_qty
        """, (rec.date, rec.store_id, rec.phone_sales, rec.ncme_sales,
              rec.phone_qty, rec.key_model_qty, rec.accessory_sales, rec.trade_in_qty))
        upserted += 1

    await db.commit()
    return {"status": "ok", "upserted": upserted}


@router.post("/targets")
async def set_monthly_targets(
    targets: List[MonthlyTarget],
    user: dict = Depends(get_current_user)):
    """批量设置月度目标（admin）"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可设置目标")

    db = await get_db()
    upserted = 0
    for t in targets:
        await db.execute("""
            INSERT INTO monthly_targets (year, month, store_id, grade,
                phone_sales_target, ncme_target, phone_qty_target,
                key_model_target, accessory_target, trade_in_rate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(year, month, store_id)
            DO UPDATE SET grade=excluded.grade,
                         phone_sales_target=excluded.phone_sales_target,
                         ncme_target=excluded.ncme_target,
                         phone_qty_target=excluded.phone_qty_target,
                         key_model_target=excluded.key_model_target,
                         accessory_target=excluded.accessory_target,
                         trade_in_rate=excluded.trade_in_rate
        """, (t.year, t.month, t.store_id, t.grade,
              t.phone_sales_target, t.ncme_target, t.phone_qty_target,
              t.key_model_target, t.accessory_target, t.trade_in_rate))
        upserted += 1

    await db.commit()
    return {"status": "ok", "upserted": upserted}


@router.get("/targets")
async def get_monthly_targets(year: int, month: int,
    user: dict = Depends(get_current_user)):
    """获取指定月份各门店的目标"""

    db = await get_db()
    cursor = await db.execute("""
        SELECT mt.*, s.name as store_name, s.province
        FROM monthly_targets mt
        JOIN stores s ON mt.store_id = s.id
        WHERE mt.year=? AND mt.month=?
        ORDER BY s.sort_order
    """, (year, month))
    return [dict(row) for row in await cursor.fetchall()]


@router.get("/latest-month")
async def get_latest_data_month(
    user: dict = Depends(get_current_user)
):
    """获取最近有销售数据的年月"""

    db = await get_db()
    cursor = await db.execute("""
        SELECT substr(date, 1, 7) as ym FROM daily_sales
        WHERE phone_sales > 0 OR phone_qty > 0 OR ncme_sales > 0
        ORDER BY date DESC LIMIT 1
    """)
    row = await cursor.fetchone()
    if row:
        parts = row["ym"].split("-")
        return {"year": int(parts[0]), "month": int(parts[1])}
    return {}


# ─── 新增：对比分析 ─────────────────────────────────────────────

@router.get("/compare")
async def compare_months(
    year: int, month: int,
    compare_type: str = "mom",  # mom=环比上月, yoy=同比去年
    user: dict = Depends(get_current_user)
):
    """多期对比分析：环比/同比"""
    db = await get_db()
    store_filter = "AND s.id=?" if (user["role"] == "staff" and user.get("store_id")) else ""
    store_params_extra = [user["store_id"]] if (user["role"] == "staff" and user.get("store_id")) else []

    def get_month_data(y, m):
        ms = f"{y}-{m:02d}-01"
        nm = f"{y}-{m + 1:02d}-01" if m < 12 else f"{y + 1}-01-01"
        sql = f"""
            SELECT s.name as store_name,
                COALESCE(SUM(ds.phone_sales),0) as phone_sales,
                COALESCE(SUM(ds.phone_qty),0) as phone_qty,
                COALESCE(SUM(ds.ncme_sales),0) as ncme_sales,
                COALESCE(SUM(ds.accessory_sales),0) as acc_sales
            FROM stores s
            LEFT JOIN daily_sales ds ON ds.store_id=s.id AND ds.date>=? AND ds.date<?
            WHERE s.is_active=1 {store_filter}
            GROUP BY s.id ORDER BY s.sort_order
        """
        params = [ms, nm] + store_params_extra
        return ms, nm, sql, params

    current = get_month_data(year, month)

    if compare_type == "mom":
        # 环比：上月
        if month == 1:
            compare = get_month_data(year - 1, 12)
        else:
            compare = get_month_data(year, month - 1)
    elif compare_type == "yoy":
        compare = get_month_data(year - 1, month)
    else:
        compare = current  # fallback

    # 查询当前月
    cur = await db.execute(current[2], tuple(current[3]))
    cur_rows = {r["store_name"]: dict(r) for r in await cur.fetchall()}

    # 查询对比月
    cmp = await db.execute(compare[2], tuple(compare[3]))
    cmp_rows = {r["store_name"]: dict(r) for r in await cmp.fetchall()}

    stores = []
    for name in cur_rows:
        c = cur_rows[name]
        p = cmp_rows.get(name, {})
        stores.append({
            "store_name": name,
            "current": {
                "phone_sales": c["phone_sales"], "phone_qty": c["phone_qty"],
                "ncme_sales": c["ncme_sales"], "acc_sales": c["acc_sales"],
            },
            "compare": {
                "phone_sales": p.get("phone_sales", 0), "phone_qty": p.get("phone_qty", 0),
                "ncme_sales": p.get("ncme_sales", 0), "acc_sales": p.get("acc_sales", 0),
            },
            "change": {
                "phone_sales_pct": round((c["phone_sales"] - p.get("phone_sales", 0)) / max(p.get("phone_sales", 1), 1) * 100, 1),
                "phone_qty_pct": round((c["phone_qty"] - p.get("phone_qty", 0)) / max(p.get("phone_qty", 1), 1) * 100, 1),
                "ncme_pct": round((c["ncme_sales"] - p.get("ncme_sales", 0)) / max(p.get("ncme_sales", 1), 1) * 100, 1),
            }
        })

    # 总计
    cur_total = {k: sum(s["current"][k] for s in stores) for k in ["phone_sales", "phone_qty", "ncme_sales", "acc_sales"]}
    cmp_total = {k: sum(s["compare"][k] for s in stores) for k in ["phone_sales", "phone_qty", "ncme_sales", "acc_sales"]}

    return {
        "compare_type": compare_type,
        "year": year, "month": month,
        "stores": stores,
        "totals": {
            "current": cur_total,
            "compare": cmp_total,
            "change_pct": {
                "phone_sales": round((cur_total["phone_sales"] - cmp_total["phone_sales"]) / max(cmp_total["phone_sales"], 1) * 100, 1),
                "phone_qty": round((cur_total["phone_qty"] - cmp_total["phone_qty"]) / max(cmp_total["phone_qty"], 1) * 100, 1),
            }
        }
    }


@router.get("/ranking")
async def get_store_ranking(
    year: int, month: int,
    metric: str = "phone_sales_rate",  # phone_sales_rate, phone_qty, ncme_sales, acc_rate
    user: dict = Depends(get_current_user)
):
    """门店排名：按指标排序"""
    db = await get_db()
    month_start = f"{year}-{month:02d}-01"
    next_month = f"{year}-{month + 1:02d}-01" if month < 12 else f"{year + 1}-01-01"

    sql = """
        SELECT s.name as store_name, s.province,
            mt.phone_sales_target, mt.ncme_target, mt.key_model_target, mt.accessory_target,
            COALESCE(SUM(ds.phone_sales),0) as phone_done,
            COALESCE(SUM(ds.phone_qty),0) as phone_qty,
            COALESCE(SUM(ds.ncme_sales),0) as ncme_done,
            COALESCE(SUM(ds.accessory_sales),0) as acc_done,
            COALESCE(SUM(ds.trade_in_qty),0) as trade_in_done
        FROM stores s
        LEFT JOIN monthly_targets mt ON mt.store_id=s.id AND mt.year=? AND mt.month=?
        LEFT JOIN daily_sales ds ON ds.store_id=s.id AND ds.date>=? AND ds.date<?
        WHERE s.is_active=1
        GROUP BY s.id
    """
    cur = await db.execute(sql, (year, month, month_start, next_month))
    rows = await cur.fetchall()

    result = []
    for r in rows:
        pt = r["phone_sales_target"] or 0
        nt = r["ncme_target"] or 0
        at = r["accessory_target"] or 0
        pd = r["phone_done"]
        nd = r["ncme_done"]
        ad = r["acc_done"]

        result.append({
            "store_name": r["store_name"],
            "province": r["province"],
            "phone_sales_rate": round(pd / pt * 100, 1) if pt > 0 else 0,
            "ncme_rate": round(nd / nt * 100, 1) if nt > 0 else 0,
            "acc_rate": round(ad / at * 100, 1) if at > 0 else 0,
            "phone_qty": r["phone_qty"],
            "phone_sales": pd,
            "ncme_sales": nd,
            "trade_in": r["trade_in_done"],
        })

    # 按指标排序
    result.sort(key=lambda x: x.get(metric, 0), reverse=True)

    # 添加排名标记
    for i, item in enumerate(result):
        item["rank"] = i + 1
        if i == 0:
            item["medal"] = "🥇"
        elif i == 1:
            item["medal"] = "🥈"
        elif i == 2:
            item["medal"] = "🥉"
        else:
            item["medal"] = ""

    return {"metric": metric, "year": year, "month": month, "stores": result}


@router.get("/export-csv")
async def export_monthly_csv(year: int, month: int,
    user: dict = Depends(get_current_user)):
    """导出月度销售进度表 CSV（方便数据分析）"""
    db = await get_db()
    month_start = f"{year}-{month:02d}-01"
    next_month = f"{year}-{month + 1:02d}-01" if month < 12 else f"{year + 1}-01-01"

    cur = await db.execute("SELECT id, name FROM stores WHERE is_active=1 ORDER BY sort_order")
    stores = await cur.fetchall()

    cur = await db.execute(
        "SELECT store_id, grade, phone_sales_target, ncme_target, phone_qty_target, key_model_target, accessory_target, trade_in_rate FROM monthly_targets WHERE year=? AND month=?",
        (year, month))
    targets = {r["store_id"]: dict(r) for r in await cur.fetchall()}

    cur = await db.execute("""
        SELECT store_id, SUM(phone_sales) as ps, SUM(ncme_sales) as ns,
               SUM(phone_qty) as pq, SUM(key_model_qty) as km,
               SUM(accessory_sales) as acs, SUM(trade_in_qty) as ti
        FROM daily_sales WHERE date>=? AND date<? GROUP BY store_id
    """, (month_start, next_month))
    sales = {r["store_id"]: dict(r) for r in await cur.fetchall()}

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["等级", "门店", "省份", "手机目标", "手机完成", "完成率%", "NCME目标", "NCME完成", "NCME率%",
                      "手机台量", "重点机型目标", "重点机型完成", "重点机型率%",
                      "配件目标", "配件完成", "配件率%", "回收台量", "回收占比%"])

    for s in stores:
        t = targets.get(s["id"], {})
        sl = sales.get(s["id"], {})
        pt = t.get("phone_sales_target", 0) or 0
        nt = t.get("ncme_target", 0) or 0
        kt = t.get("key_model_target", 0) or 0
        at = t.get("accessory_target", 0) or 0
        pd = sl.get("ps", 0) or 0
        nd = sl.get("ns", 0) or 0
        kd = sl.get("km", 0) or 0
        ad = sl.get("acs", 0) or 0
        pq = sl.get("pq", 0) or 0
        ti = sl.get("ti", 0) or 0

        writer.writerow([
            t.get("grade", ""), s["name"], "",
            pt, pd, round(pd / pt * 100, 1) if pt > 0 else 0,
            nt, nd, round(nd / nt * 100, 1) if nt > 0 else 0,
            pq, kt, kd, round(kd / kt * 100, 1) if kt > 0 else 0,
            at, ad, round(ad / at * 100, 1) if at > 0 else 0,
            ti, round(ti / pq * 100, 1) if pq > 0 else 0,
        ])

    output.seek(0)
    from urllib.parse import quote
    filename = f"销售数据_{year}年{month}月.csv"
    return StreamingResponse(
        iter([output.getvalue().encode('utf-8-sig')]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@router.get("/summary")
async def get_sales_summary(
    user: dict = Depends(get_current_user)
):

    db = await get_db()
    today = __import__('datetime').datetime.now().strftime('%Y-%m-%d')
    month_start = today[:7] + "-01"
    next_month_date = __import__('datetime').date(
        int(today[:4]), int(today[5:7]) + 1, 1) if int(today[5:7]) < 12 \
        else __import__('datetime').date(int(today[:4]) + 1, 1, 1)
    next_month = str(next_month_date)

    # 今日合计
    cursor = await db.execute("""
        SELECT SUM(phone_sales) as today_sales, SUM(phone_qty) as today_qty,
               SUM(ncme_sales) as today_ncme, SUM(accessory_sales) as today_acc,
               SUM(trade_in_qty) as today_trade_in
        FROM daily_sales WHERE date = ?
    """, (today,))
    today_data = await cursor.fetchone()

    # 本月合计
    cursor = await db.execute("""
        SELECT SUM(phone_sales) as month_sales, SUM(phone_qty) as month_qty,
               SUM(ncme_sales) as month_ncme, SUM(accessory_sales) as month_acc,
               SUM(trade_in_qty) as month_trade_in
        FROM daily_sales WHERE date >= ? AND date < ?
    """, (month_start, next_month))
    month_data = await cursor.fetchone()

    return {
        "today": {
            "sales": today_data["today_sales"] or 0,
            "qty": today_data["today_qty"] or 0,
            "ncme": today_data["today_ncme"] or 0,
            "accessory": today_data["today_acc"] or 0,
            "trade_in": today_data["today_trade_in"] or 0,
        },
        "month": {
            "sales": month_data["month_sales"] or 0,
            "qty": month_data["month_qty"] or 0,
            "ncme": month_data["month_ncme"] or 0,
            "accessory": month_data["month_acc"] or 0,
            "trade_in": month_data["month_trade_in"] or 0,
        }
    }


@router.get("/export")
async def export_monthly_excel(year: int, month: int,
    user: dict = Depends(get_current_user)):
    """导出月度销售进度表Excel（从数据库重新生成，和桌面版格式一致）"""

    db = await get_db()
    month_start = f"{year}-{month:02d}-01"
    next_month = f"{year}-{month + 1:02d}-01" if month < 12 else f"{year + 1}-01-01"

    # 获取所有活跃门店（只查一次）
    cursor = await db.execute("SELECT id, name FROM stores WHERE is_active=1 ORDER BY sort_order")
    stores = await cursor.fetchall()

    # 2. 一次查出全部月度目标
    cursor = await db.execute(
        "SELECT store_id, grade, phone_sales_target, ncme_target, phone_qty_target, key_model_target, accessory_target, trade_in_rate FROM monthly_targets WHERE year=? AND month=?",
        (year, month))
    targets_map = {r["store_id"]: dict(r) for r in await cursor.fetchall()}

    # 3. 一次查出全部销售汇总
    cursor = await db.execute("""
        SELECT store_id,
            SUM(phone_sales) as total_phone_sales,
            SUM(ncme_sales) as total_ncme_sales,
            SUM(phone_qty) as total_phone_qty,
            SUM(key_model_qty) as total_key_model_qty,
            SUM(accessory_sales) as total_accessory_sales,
            SUM(trade_in_qty) as total_trade_in_qty
        FROM daily_sales WHERE date >= ? AND date < ?
        GROUP BY store_id
    """, (month_start, next_month))
    sales_map = {r["store_id"]: r for r in await cursor.fetchall()}

    # 4. 组装数据
    rows_data = []
    totals = {
        "phone_target": 0, "phone_done": 0,
        "ncme_target": 0, "ncme_done": 0,
        "phone_qty": 0, "km_target": 0, "km_done": 0,
        "acc_target": 0, "acc_done": 0,
        "trade_in": 0
    }

    for store in stores:
        sid = store["id"]
        target = targets_map.get(sid, {})
        sales = sales_map.get(sid)

        phone_done = sales["total_phone_sales"] or 0 if sales else 0
        phone_target = target.get("phone_sales_target", 0)
        ncme_done = sales["total_ncme_sales"] or 0 if sales else 0
        ncme_target = target.get("ncme_target", 0)
        phone_qty = sales["total_phone_qty"] or 0 if sales else 0
        km_done = sales["total_key_model_qty"] or 0 if sales else 0
        km_target = target.get("key_model_target", 0)
        acc_done = sales["total_accessory_sales"] or 0 if sales else 0
        acc_target = target.get("accessory_target", 0)
        trade_in = sales["total_trade_in_qty"] or 0 if sales else 0

        totals["phone_target"] += phone_target
        totals["phone_done"] += phone_done
        totals["ncme_target"] += ncme_target
        totals["ncme_done"] += ncme_done
        totals["phone_qty"] += phone_qty
        totals["km_target"] += km_target
        totals["km_done"] += km_done
        totals["acc_target"] += acc_target
        totals["acc_done"] += acc_done
        totals["trade_in"] += trade_in

        rows_data.append({
            "store_name": store["name"],
            "grade": target.get("grade", ""),
            "phone_target": phone_target, "phone_done": phone_done,
            "ncme_target": ncme_target, "ncme_done": ncme_done,
            "phone_qty": phone_qty,
            "km_target": km_target, "km_done": km_done,
            "acc_target": acc_target, "acc_done": acc_done,
            "trade_in": trade_in,
        })


    # 3. 生成Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月销售进度"

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # 标题行
    headers = [
        ('等级', 6), ('门店', 22),
        ('手机销售目标', 14), ('完成', 14), ('完成率', 10),
        ('NCME目标', 14), ('完成', 14), ('完成率', 10),
        ('手机台量', 10),
        ('重点机型目标', 12), ('完成', 10), ('完成率', 10),
        ('配件目标', 14), ('完成', 14), ('完成率', 10),
        ('回收', 10), ('回收占比', 10)
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 数据行
    for i, rd in enumerate(rows_data):
        row = i + 2
        phone_rate = rd["phone_done"] / rd["phone_target"] if rd["phone_target"] > 0 else 0
        ncme_rate = rd["ncme_done"] / rd["ncme_target"] if rd["ncme_target"] > 0 else 0
        km_rate = rd["km_done"] / rd["km_target"] if rd["km_target"] > 0 else 0
        acc_rate = rd["acc_done"] / rd["acc_target"] if rd["acc_target"] > 0 else 0
        ti_rate = rd["trade_in"] / rd["phone_qty"] if rd["phone_qty"] > 0 else 0

        values = [
            rd["grade"], rd["store_name"],
            rd["phone_target"], rd["phone_done"], phone_rate,
            rd["ncme_target"], rd["ncme_done"], ncme_rate,
            rd["phone_qty"],
            rd["km_target"], rd["km_done"], km_rate,
            rd["acc_target"], rd["acc_done"], acc_rate,
            rd["trade_in"], ti_rate
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center if col_idx in (1, 5, 8, 9, 12, 15, 16, 17) else right_align
            # 完成率列设百分比格式
            if col_idx in (5, 8, 12, 15, 17):
                cell.number_format = '0.0%'

    # 合计行
    total_row = len(rows_data) + 2
    total_phone_rate = totals["phone_done"] / totals["phone_target"] if totals["phone_target"] > 0 else 0
    total_ncme_rate = totals["ncme_done"] / totals["ncme_target"] if totals["ncme_target"] > 0 else 0
    total_km_rate = totals["km_done"] / totals["km_target"] if totals["km_target"] > 0 else 0
    total_acc_rate = totals["acc_done"] / totals["acc_target"] if totals["acc_target"] > 0 else 0
    total_ti_rate = totals["trade_in"] / totals["phone_qty"] if totals["phone_qty"] > 0 else 0

    total_values = [
        '', '合计',
        totals["phone_target"], totals["phone_done"], total_phone_rate,
        totals["ncme_target"], totals["ncme_done"], total_ncme_rate,
        totals["phone_qty"],
        totals["km_target"], totals["km_done"], total_km_rate,
        totals["acc_target"], totals["acc_done"], total_acc_rate,
        totals["trade_in"], total_ti_rate
    ]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(name='微软雅黑', bold=True, size=11)
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center if col_idx in (1, 5, 8, 9, 12, 15, 16, 17) else right_align
        if col_idx in (5, 8, 12, 15, 17):
            cell.number_format = '0.0%'

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"月度销售进度表_{year}年{month}月.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/export-daily")
async def export_daily_excel(date: str,
    user: dict = Depends(get_current_user)):
    """导出指定日期的销售明细表Excel（只导出当天数据，避免重复计算）"""
    import re
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        raise HTTPException(status_code=400, detail="日期格式错误，应为 YYYY-MM-DD")

    db = await get_db()

    # 获取所有活跃门店
    cursor = await db.execute("SELECT id, name FROM stores WHERE is_active=1 ORDER BY sort_order")
    stores = await cursor.fetchall()

    # 查询指定日期的销售数据
    cursor = await db.execute("""
        SELECT store_id,
            phone_sales as total_phone_sales,
            ncme_sales as total_ncme_sales,
            phone_qty as total_phone_qty,
            key_model_qty as total_key_model_qty,
            accessory_sales as total_accessory_sales,
            trade_in_qty as total_trade_in_qty
        FROM daily_sales WHERE date = ?
    """, (date,))
    sales_map = {r["store_id"]: r for r in await cursor.fetchall()}

    # 查询指定日期的月度目标（用于计算完成率）
    month_start = date[:7] + "-01"
    next_month = f"{date[:4]}-{int(date[5:7]) + 1:02d}-01" if int(date[5:7]) < 12 else f"{int(date[:4]) + 1}-01-01"
    cursor = await db.execute(
        "SELECT store_id, grade, phone_sales_target, ncme_target, phone_qty_target, key_model_target, accessory_target, trade_in_rate FROM monthly_targets WHERE year=? AND month=?",
        (int(date[:4]), int(date[5:7])))
    targets_map = {r["store_id"]: dict(r) for r in await cursor.fetchall()}

    # 组装数据
    rows_data = []
    totals = {
        "phone_target": 0, "phone_done": 0,
        "ncme_target": 0, "ncme_done": 0,
        "phone_qty": 0, "km_target": 0, "km_done": 0,
        "acc_target": 0, "acc_done": 0,
        "trade_in": 0
    }

    for store in stores:
        sid = store["id"]
        target = targets_map.get(sid, {})
        sales = sales_map.get(sid)

        phone_done = sales["total_phone_sales"] or 0 if sales else 0
        phone_target = target.get("phone_sales_target", 0)
        ncme_done = sales["total_ncme_sales"] or 0 if sales else 0
        ncme_target = target.get("ncme_target", 0)
        phone_qty = sales["total_phone_qty"] or 0 if sales else 0
        km_done = sales["total_key_model_qty"] or 0 if sales else 0
        km_target = target.get("key_model_target", 0)
        acc_done = sales["total_accessory_sales"] or 0 if sales else 0
        acc_target = target.get("accessory_target", 0)
        trade_in = sales["total_trade_in_qty"] or 0 if sales else 0

        totals["phone_target"] += phone_target or 0
        totals["phone_done"] += phone_done
        totals["ncme_target"] += ncme_target or 0
        totals["ncme_done"] += ncme_done
        totals["phone_qty"] += phone_qty
        totals["km_target"] += km_target or 0
        totals["km_done"] += km_done
        totals["acc_target"] += acc_target or 0
        totals["acc_done"] += acc_done
        totals["trade_in"] += trade_in

        rows_data.append({
            "store_name": store["name"],
            "grade": target.get("grade", ""),
            "phone_target": phone_target or 0,
            "phone_done": phone_done,
            "ncme_target": ncme_target or 0,
            "ncme_done": ncme_done,
            "phone_qty": phone_qty,
            "km_target": km_target or 0,
            "km_done": km_done,
            "acc_target": acc_target or 0,
            "acc_done": acc_done,
            "trade_in": trade_in,
        })

    # 生成Excel（复用月度导出的格式）
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{date} 销售明细"

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # 标题行
    ws.row_dimensions[1].height = 30
    headers = [
        ("等级", 6), ("门店", 22),
        ("手机目标", 14), ("完成", 14), ("完成率%", 10),
        ("NCME目标", 14), ("完成", 14), ("完成率%", 10),
        ("手机台量", 10),
        ("重点机型目标", 12), ("完成", 10), ("完成率%", 10),
        ("配件目标", 14), ("完成", 14), ("完成率%", 10),
        ("回收台量", 10), ("回收占比%", 10)
    ]
    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 数据行
    for i, rd in enumerate(rows_data, 2):
        phone_rate = rd["phone_done"] / rd["phone_target"] if rd["phone_target"] > 0 else 0
        ncme_rate = rd["ncme_done"] / rd["ncme_target"] if rd["ncme_target"] > 0 else 0
        km_rate = rd["km_done"] / rd["km_target"] if rd["km_target"] > 0 else 0
        acc_rate = rd["acc_done"] / rd["acc_target"] if rd["acc_target"] > 0 else 0
        ti_rate = rd["trade_in"] / rd["phone_qty"] if rd["phone_qty"] > 0 else 0

        values = [
            rd["grade"], rd["store_name"],
            rd["phone_target"], rd["phone_done"], phone_rate,
            rd["ncme_target"], rd["ncme_done"], ncme_rate,
            rd["phone_qty"],
            rd["km_target"], rd["km_done"], km_rate,
            rd["acc_target"], rd["acc_done"], acc_rate,
            rd["trade_in"], ti_rate
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center if col_idx in (1, 5, 8, 9, 12, 15, 17) else right_align
            if col_idx in (5, 8, 12, 15, 17):
                cell.number_format = '0.0%'

    # 合计行
    total_row = len(rows_data) + 2
    total_phone_rate = totals["phone_done"] / totals["phone_target"] if totals["phone_target"] > 0 else 0
    total_ncme_rate = totals["ncme_done"] / totals["ncme_target"] if totals["ncme_target"] > 0 else 0
    total_km_rate = totals["km_done"] / totals["km_target"] if totals["km_target"] > 0 else 0
    total_acc_rate = totals["acc_done"] / totals["acc_target"] if totals["acc_target"] > 0 else 0
    total_ti_rate = totals["trade_in"] / totals["phone_qty"] if totals["phone_qty"] > 0 else 0

    total_values = [
        '', '合计',
        totals["phone_target"], totals["phone_done"], total_phone_rate,
        totals["ncme_target"], totals["ncme_done"], total_ncme_rate,
        totals["phone_qty"],
        totals["km_target"], totals["km_done"], total_km_rate,
        totals["acc_target"], totals["acc_done"], total_acc_rate,
        totals["trade_in"], total_ti_rate
    ]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(name='微软雅黑', bold=True, size=11)
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center if col_idx in (1, 5, 8, 9, 12, 15, 17) else right_align
        if col_idx in (5, 8, 12, 15, 17):
            cell.number_format = '0.0%'

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"销售明细_{date}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
