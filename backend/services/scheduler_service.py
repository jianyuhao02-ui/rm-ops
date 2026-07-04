"""
定时任务调度器
使用 APScheduler 管理定时任务：
- 每小时 eBoss目录扫描 → 零售汇总/库存汇总 → 写入数据库 → 更新Excel → 推送通知
- 每周一/四9:30 库存预警检查 → 推送通知
- 每日10:00 价格异常检查 → 推送通知
- 每日凌晨3:00 数据库备份
"""
import os
import re
import shutil
import asyncio
import aiosqlite
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from backend.logger import scheduler_logger
from backend.config import (
    DB_PATH, DB_BACKUP_DIR, EBOSS_SCAN_DIR,
    INVENTORY_ALERT_DAYS, INVENTORY_ALERT_HOUR, INVENTORY_ALERT_MINUTE,
    PRICE_CHECK_HOUR, BACKUP_HOUR, EBOSS_SCAN_INTERVAL_MINUTES
)

logger = scheduler_logger

# 路径常量
DB_PATH_STR = str(DB_PATH)
BACKUP_DIR = str(DB_BACKUP_DIR)
PROJECT_DIR = str(Path(__file__).parent.parent.parent)

# 模型名称映射（数据库编码 → 中文名）
MODEL_NAMES = {
    "S9420": "Galaxy S26",
    "S9470": "Galaxy S26+",
    "S9480": "Galaxy S26 Ultra",
    "ZFOLD7": "Galaxy Z Fold7",
    "ZFLIP7": "Galaxy Z Flip7",
    "W26": "Galaxy W26",
}

scheduler = AsyncIOScheduler()


# ==================== 通用工具 ====================

async def _notify_admin(title: str, content: str):
    """通知管理员（写数据库通知 + 钉钉私信）"""
    db = await aiosqlite.connect(DB_PATH_STR)
    db.row_factory = aiosqlite.Row
    try:
        await db.execute(
            "INSERT INTO notifications (title, content, type) VALUES (?,?,?)",
            (title, content, "system")
        )
        await db.commit()
    finally:
        await db.close()

    try:
        proc = await asyncio.create_subprocess_exec(
            "dws", "chat", "message", "send-direct",
            "--user", "16381712592737652",
            "--message", f"【{title}】\n{content[:500]}",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        await proc.wait()
        if proc.returncode == 0:
            logger.info(f"钉钉私信发送成功: {title}")
        else:
            err = await proc.stderr.read()
            logger.warning(f"钉钉私信发送失败: {err.decode('utf-8', errors='ignore')[:200]}")
    except Exception as e:
        logger.error(f"钉钉私信异常: {e}")


def _get_store_name_by_id(store_id: int) -> str:
    """store_id → 门店名"""
    names = {
        1: "万象城", 2: "万象汇", 3: "兴义", 4: "遵义",
        5: "曲靖", 6: "六盘水", 7: "龙湾", 8: "昭通", 9: "清镇", 10: "安顺"
    }
    return names.get(store_id, "未知")


# ==================== 库存预警任务 ====================

async def task_inventory_alert():
    """
    每周一/四9:30执行
    检查库存预警规则 → 生成预警通知 → 钉钉推送
    """
    logger.info("=== 开始库存预警检查 ===")
    try:
        alerts = await _get_inventory_alerts()
        if not alerts:
            logger.info("无库存预警")
            return

        from backend.services.dingtalk_service import build_inventory_alert_md
        alert_md = build_inventory_alert_md(alerts)
        await _notify_admin("库存预警", alert_md)
        logger.info(f"库存预警完成: {len(alerts)} 条预警")
    except Exception as e:
        logger.error(f"库存预警失败: {e}", exc_info=True)


async def _get_inventory_alerts() -> list:
    """获取库存预警列表"""
    db = await aiosqlite.connect(DB_PATH_STR)
    db.row_factory = aiosqlite.Row
    try:
        cursor = await db.execute("SELECT * FROM alert_rules WHERE is_active=1")
        rules = await cursor.fetchall()
        rule_map = {r["model_series"]: dict(r) for r in rules}

        cursor = await db.execute("""
            SELECT i.*, s.name as store_name
            FROM inventory i JOIN stores s ON i.store_id = s.id
            WHERE s.is_active = 1
        """)
        all_inv = await cursor.fetchall()

        alerts = []

        def get_series(mc):
            if not mc:
                return ""
            s = mc.upper()
            if any(k in s for k in ("S9420", "S9470", "S9480")):
                return "S26"
            if "ZFOLD7" in s:
                return "FOLD7"
            if "ZFLIP7" in s:
                return "FLIP7"
            if "W26" in s:
                return "W26"
            return ""

        if "S26" in rule_map and rule_map["S26"]["rule_type"] == "per_store_color_spec":
            threshold = rule_map["S26"]["threshold"]
            for inv in all_inv:
                if get_series(inv["model_code"]) != "S26":
                    continue
                if inv["qty"] < threshold:
                    alerts.append({
                        "series": "S26",
                        "level": "danger" if inv["qty"] == 0 else "warning",
                        "message": f"{inv['store_name']} {inv['model_code']} {inv['color']} {inv['spec']} 仅{inv['qty']}台（需≥{threshold}台）"
                    })

        if "FOLD7" in rule_map:
            threshold = rule_map["FOLD7"]["threshold"]
            st = defaultdict(int)
            for i in all_inv:
                if get_series(i["model_code"]) == "FOLD7":
                    st[i["store_name"]] += i["qty"]
            for sn, total in st.items():
                if total < threshold:
                    alerts.append({
                        "series": "FOLD7",
                        "level": "danger" if total == 0 else "warning",
                        "message": f"{sn} Z Fold7 合计仅{total}台（需≥{threshold}台）"
                    })

        if "FLIP7" in rule_map:
            threshold = rule_map["FLIP7"]["threshold"]
            st = defaultdict(int)
            for i in all_inv:
                if get_series(i["model_code"]) == "FLIP7":
                    st[i["store_name"]] += i["qty"]
            for sn, total in st.items():
                if total < threshold:
                    alerts.append({
                        "series": "FLIP7",
                        "level": "danger" if total == 0 else "warning",
                        "message": f"{sn} Z Flip7 合计仅{total}台（需≥{threshold}台）"
                    })

        if "W26" in rule_map:
            threshold = rule_map["W26"]["threshold"]
            w26_total = sum(i["qty"] for i in all_inv if get_series(i["model_code"]) == "W26")
            if w26_total < threshold:
                alerts.append({
                    "series": "W26",
                    "level": "danger" if w26_total == 0 else "warning",
                    "message": f"W26 全渠道仅{w26_total}台（需≥{threshold}台）"
                })

        alerts.sort(key=lambda x: (0 if x["level"] == "danger" else 1, x["series"]))
        return alerts
    finally:
        await db.close()


# ==================== 价格检查任务 ====================

async def task_price_check():
    """
    每日10:00执行
    检查九机网价格是否有变动 → 如有变动推送通知
    """
    logger.info("=== 开始价格检查 ===")
    try:
        import httpx

        db = await aiosqlite.connect(DB_PATH_STR)
        db.row_factory = aiosqlite.Row
        try:
            cursor = await db.execute(
                "SELECT model_code, spec, price as old_price FROM price_records WHERE platform='九机网'"
            )
            records = await cursor.fetchall()
        finally:
            await db.close()

        if not records:
            logger.info("无九机网价格记录，跳过")
            return

        PRODUCT_IDS = {
            ("S9420", "12+256G"): "492798",
            ("S9470", "12+256G"): "492820",
            ("S9470", "12+512G"): "492821",
            ("S9480", "12+256G"): "492822",
            ("S9480", "12+512G"): "492823",
            ("S9480", "16+1TB"): "492824",
            ("ZFOLD7", "12+256G"): "478720",
            ("ZFOLD7", "12+512G"): "478721",
            ("ZFLIP7", "12+256G"): "478722",
            ("ZFLIP7", "12+512G"): "478723",
            ("W26", "16+512G"): "478725",
            ("W26", "16+1TB"): "478724",
        }

        changes = []
        async with httpx.AsyncClient(timeout=15) as client:
            for r in records:
                key = (r["model_code"], r["spec"])
                pid = PRODUCT_IDS.get(key)
                if not pid:
                    continue

                try:
                    url = f"https://www.jiuji.com/product/{pid}.html"
                    resp = await client.get(url, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True)
                    price_match = re.search(r'"price"\s*:\s*(\d+)', resp.text)
                    if not price_match:
                        continue
                    new_price = int(price_match.group(1))
                    old_price = int(r["old_price"]) if r["old_price"] else 0

                    if old_price > 0 and new_price != old_price:
                        changes.append({
                            "model_name": MODEL_NAMES.get(r["model_code"], r["model_code"]),
                            "spec": r["spec"],
                            "platform": "九机网",
                            "old_price": old_price,
                            "new_price": new_price,
                            "diff": new_price - old_price,
                        })
                        db2 = await aiosqlite.connect(DB_PATH_STR)
                        try:
                            await db2.execute(
                                "UPDATE price_records SET price=? WHERE model_code=? AND spec=? AND platform='九机网'",
                                (new_price, r["model_code"], r["spec"])
                            )
                            await db2.commit()
                        finally:
                            await db2.close()
                except Exception as e:
                    logger.warning(f"抓取 {key} 价格失败: {e}")
                    continue

        if changes:
            from backend.services.dingtalk_service import build_price_alert_md
            alert_md = build_price_alert_md(changes)
            await _notify_admin("价格变动", alert_md)
            logger.info(f"价格检查完成: {len(changes)} 项变动")
        else:
            logger.info("价格检查完成: 无变动")

    except Exception as e:
        logger.error(f"价格检查失败: {e}", exc_info=True)


# ==================== 数据库备份任务 ====================

async def task_backup_database():
    """
    每日凌晨3:00执行
    备份 SQLite 数据库到桌面
    """
    logger.info("=== 开始数据库备份 ===")
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"samsung_ops_{timestamp}.db"
        backup_path = os.path.join(BACKUP_DIR, backup_name)

        if os.path.exists(DB_PATH_STR):
            shutil.copy2(DB_PATH_STR, backup_path)
            logger.info(f"数据库已备份: {backup_path}")

            cutoff = datetime.now().timestamp() - 7 * 24 * 3600
            for f in Path(BACKUP_DIR).glob("samsung_ops_*.db"):
                if f.stat().st_mtime < cutoff:
                    f.unlink()
                    logger.info(f"清理旧备份: {f.name}")
        else:
            logger.warning(f"数据库文件不存在: {DB_PATH}")
    except Exception as e:
        logger.error(f"数据库备份失败: {e}", exc_info=True)


# ==================== eBoss 目录扫描任务 ====================

async def task_scan_eboss_dir():
    r"""
    每小时执行
    扫描 C:\eBoss\Local\ 目录，自动识别新文件并导入：
    - 零售汇总 → daily_sales 表 → 更新桌面Excel
    - 库存汇总 → inventory 表
    """
    logger.info("=== 开始 eBoss 目录扫描 ===")
    try:
        from backend.services.eboss_parser import (
            scan_eboss_directory, parse_retail_summary, parse_retail_detail,
            parse_inventory_xls, parse_inventory_xlsx, parse_trade_in, match_store_id
        )

        files = scan_eboss_directory()
        if not files:
            logger.info("eBoss 目录无新文件")
            return

        processed = await _get_processed_eboss_files()
        new_files = [f for f in files if f["filename"] not in processed]

        if not new_files:
            logger.info(f"eBoss 目录 {len(files)} 个文件已全部处理过")
            return

        logger.info(f"eBoss 发现 {len(new_files)} 个新文件待处理")

        imported_sales = 0
        imported_inv = 0
        imported_dates = set()  # 记录导入了哪些日期的数据

        for f in new_files:
            fname = f["filename"]
            ftype = f["filetype"]
            fpath = f["filepath"]
            mtime = f["mtime"]

            try:
                if "retail_detail" in ftype:
                    records = parse_retail_detail(fpath)
                    # 零售明细是最准确的格式，替换同期 eboss 数据
                    import_dates = list(set(r["date"] for r in records))
                    count = await _import_eboss_sales_replace(records, import_dates)
                    imported_sales += count
                    imported_dates.update(import_dates)
                    await _save_eboss_sync_log(fname, "retail_detail", None, mtime,
                                              len(records), count, "success")

                elif "retail_summary" in ftype:
                    records = parse_retail_summary(fpath)
                    count = await _import_eboss_sales(records)
                    imported_sales += count
                    imported_dates.update(r["date"] for r in records)
                    await _save_eboss_sync_log(fname, "retail_summary", None, mtime,
                                              len(records), count, "success")

                elif ftype in ("inventory", "inventory_xlsx"):
                    if ftype == "inventory":
                        records = parse_inventory_xls(fpath)
                    else:
                        records = parse_inventory_xlsx(fpath)
                    count = await _import_eboss_inventory(records)
                    imported_inv += count
                    await _save_eboss_sync_log(fname, "inventory", None, mtime,
                                              len(records), count, "success")

                elif ftype == "retail_order":
                    await _save_eboss_sync_log(fname, "retail_order", None, mtime,
                                              0, 0, "skipped", "零售单暂不自动导入")

                elif ftype == "trade_in":
                    trade_records = parse_trade_in(fpath)
                    if trade_records:
                        await _import_trade_in(trade_records)
                        imported_dates.update(r["date"] for r in trade_records)
                    await _save_eboss_sync_log(fname, "trade_in", None, mtime,
                                              len(trade_records) if trade_records else 0,
                                              len(trade_records) if trade_records else 0,
                                              "success")

                logger.info(f"eBoss 文件处理完成: {fname} ({ftype})")

            except Exception as e:
                logger.error(f"eBoss 文件处理失败: {fname}: {e}", exc_info=True)
                await _save_eboss_sync_log(fname, ftype, None, mtime,
                                          0, 0, "error", str(e)[:500])

        # 有销售数据导入时，从DB重新生成桌面Excel（避免重复计算）
        if imported_sales > 0 and imported_dates:
            await _regenerate_excel_from_db(imported_dates)

        # 汇总通知
        if imported_sales > 0 or imported_inv > 0:
            msg = f"eBoss自动同步完成：\n- 销售数据 {imported_sales} 条\n- 库存数据 {imported_inv} 条"
            await _notify_admin("eBoss同步", msg)
            logger.info(f"eBoss 同步完成: 销售={imported_sales}, 库存={imported_inv}")
        else:
            logger.info("eBoss 同步完成: 无新数据导入")

    except Exception as e:
        logger.error(f"eBoss 扫描任务失败: {e}", exc_info=True)


async def _regenerate_excel_from_db(imported_dates: set):
    """从数据库重新生成桌面Excel进度表（避免增量累加的重复计算问题）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # 按月份分组
        months = set()
        for d in imported_dates:
            try:
                dt = datetime.strptime(d, "%Y-%m-%d")
                months.add((dt.year, dt.month))
            except (ValueError, TypeError):
                continue

        for year, month in months:
            month_str = f"{year}年{month}月"
            excel_path = os.path.join(os.path.expanduser("~"), "Desktop", f"月度销售进度表_{month_str}.xlsx")
            month_prefix = f"{year}-{month:02d}"

            db = await aiosqlite.connect(DB_PATH_STR)
            db.row_factory = aiosqlite.Row
            try:
                cursor = await db.execute("SELECT id, name FROM stores WHERE is_active=1 ORDER BY sort_order")
                stores = await cursor.fetchall()

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"{month_str}销售进度"

                # 表头
                headers = [
                    ('等级', 6), ('门店', 22),
                    ('手机销售目标', 14), ('完成', 14), ('完成率', 10),
                    ('NCME目标', 14), ('完成', 14), ('完成率', 10),
                    ('手机台量', 10),
                    ('重点机型目标', 12), ('完成', 10), ('完成率', 10),
                    ('配件目标', 14), ('完成', 14), ('完成率', 10),
                    ('回收', 10), ('回收占比', 10)
                ]
                h_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                h_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
                center = Alignment(horizontal='center', vertical='center')
                right_a = Alignment(horizontal='right', vertical='center')
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

                ws.row_dimensions[1].height = 30
                for ci, (name, w) in enumerate(headers, 1):
                    c = ws.cell(row=1, column=ci, value=name)
                    c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = border
                    ws.column_dimensions[get_column_letter(ci)].width = w

                totals = {k: 0 for k in ['pt','pd','nt','nd','qty','kt','kd','at','ad','ti']}

                for i, store in enumerate(stores):
                    sid = store["id"]
                    row = i + 2

                    # 目标
                    cursor = await db.execute(
                        "SELECT * FROM monthly_targets WHERE year=? AND month=? AND store_id=?",
                        (year, month, sid))
                    t_row = await cursor.fetchone()
                    target = dict(t_row) if t_row else {}

                    # 销售
                    cursor = await db.execute("""
                        SELECT SUM(phone_sales), SUM(ncme_sales), SUM(phone_qty),
                               SUM(key_model_qty), SUM(accessory_sales), SUM(trade_in_qty)
                        FROM daily_sales WHERE store_id=? AND date LIKE ?
                    """, (sid, month_prefix + '%'))
                    sales = await cursor.fetchone()

                    pd = sales[0] or 0; pt = target.get("phone_sales_target", 0)
                    nd = sales[1] or 0; nt = target.get("ncme_target", 0)
                    qty = sales[2] or 0
                    kd = sales[3] or 0; kt = target.get("key_model_target", 0)
                    ad = sales[4] or 0; at = target.get("accessory_target", 0)
                    ti = sales[5] or 0

                    totals['pt']+=pt; totals['pd']+=pd; totals['nt']+=nt; totals['nd']+=nd
                    totals['qty']+=qty; totals['kt']+=kt; totals['kd']+=kd
                    totals['at']+=at; totals['ad']+=ad; totals['ti']+=ti

                    vals = [
                        target.get("grade", ""), store["name"],
                        pt, pd, pd/pt if pt > 0 else 0,
                        nt, nd, nd/nt if nt > 0 else 0,
                        qty, kt, kd, kd/kt if kt > 0 else 0,
                        at, ad, ad/at if at > 0 else 0,
                        ti, ti/qty if qty > 0 else 0
                    ]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=row, column=ci, value=v)
                        c.border = border
                        c.alignment = center if ci in (1,5,8,9,12,15,16,17) else right_a
                        if ci in (5,8,12,15,17): c.number_format = '0.0%'

                # 合计行
                tr = len(stores) + 2
                t_vals = [
                    '', '合计',
                    totals['pt'], totals['pd'], totals['pd']/totals['pt'] if totals['pt'] else 0,
                    totals['nt'], totals['nd'], totals['nd']/totals['nt'] if totals['nt'] else 0,
                    totals['qty'], totals['kt'], totals['kd'], totals['kd']/totals['kt'] if totals['kt'] else 0,
                    totals['at'], totals['ad'], totals['ad']/totals['at'] if totals['at'] else 0,
                    totals['ti'], totals['ti']/totals['qty'] if totals['qty'] else 0
                ]
                t_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                t_font = Font(name='微软雅黑', bold=True, size=11)
                for ci, v in enumerate(t_vals, 1):
                    c = ws.cell(row=tr, column=ci, value=v)
                    c.font = t_font; c.fill = t_fill; c.border = border
                    c.alignment = center if ci in (1,5,8,9,12,15,16,17) else right_a
                    if ci in (5,8,12,15,17): c.number_format = '0.0%'

                wb.save(excel_path)
                logger.info(f"Excel已重新生成: {excel_path}")
            finally:
                await db.close()

    except Exception as e:
        logger.error(f"重新生成Excel失败: {e}", exc_info=True)


async def _get_processed_eboss_files() -> set:
    """获取已处理的 eBoss 文件名集合"""
    db = await aiosqlite.connect(DB_PATH_STR)
    db.row_factory = aiosqlite.Row
    try:
        cursor = await db.execute("SELECT file_name FROM eboss_sync_log WHERE status='success'")
        rows = await cursor.fetchall()
        return {r["file_name"] for r in rows}
    finally:
        await db.close()


async def _import_eboss_sales(records: list) -> int:
    """将零售汇总解析结果写入 daily_sales 表（eBoss 来源）"""
    if not records:
        return 0
    db = await aiosqlite.connect(DB_PATH_STR)
    try:
        imported = 0
        for r in records:
            await db.execute("""
                INSERT INTO daily_sales (date, store_id, phone_sales, ncme_sales, phone_qty,
                                        key_model_qty, accessory_sales, trade_in_qty, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'eboss')
                ON CONFLICT(date, store_id) DO UPDATE SET
                    phone_sales=excluded.phone_sales,
                    ncme_sales=excluded.ncme_sales,
                    phone_qty=excluded.phone_qty,
                    key_model_qty=excluded.key_model_qty,
                    accessory_sales=excluded.accessory_sales,
                    trade_in_qty=excluded.trade_in_qty,
                    source='eboss'
            """, (r["date"], r["store_id"],
                  r.get("phone_sales", 0), r.get("ncme_sales", 0),
                  r.get("phone_qty", 0), r.get("key_model_qty", 0),
                  r.get("accessory_sales", 0), r.get("trade_in_qty", 0)))
            imported += 1
        await db.commit()
        return imported
    finally:
        await db.close()


async def _import_eboss_sales_replace(records: list, import_dates: list) -> int:
    """将零售明细解析结果写入 daily_sales 表，替换同期 eboss 数据（更准确的格式）"""
    if not records:
        return 0
    db = await aiosqlite.connect(DB_PATH_STR)
    try:
        # 先删除这些日期的 eboss 源数据
        placeholders = ",".join("?" * len(import_dates))
        cursor = await db.execute(
            f"SELECT COUNT(*) FROM daily_sales WHERE date IN ({placeholders}) AND source='eboss'",
            import_dates
        )
        deleted_count = (await cursor.fetchone())[0]

        if deleted_count > 0:
            await db.execute(
                f"DELETE FROM daily_sales WHERE date IN ({placeholders}) AND source='eboss'",
                import_dates
            )
            logger.info(f"零售明细导入: 替换 {deleted_count} 条旧 eboss 数据（{import_dates[0]}~{import_dates[-1]}）")

        # 插入新数据
        imported = 0
        for r in records:
            await db.execute("""
                INSERT OR REPLACE INTO daily_sales (date, store_id, phone_sales, ncme_sales, phone_qty,
                                                    key_model_qty, accessory_sales, trade_in_qty, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'eboss')
            """, (r["date"], r["store_id"],
                  r.get("phone_sales", 0), r.get("ncme_sales", 0),
                  r.get("phone_qty", 0), r.get("key_model_qty", 0),
                  r.get("accessory_sales", 0), r.get("trade_in_qty", 0)))
            imported += 1
        await db.commit()
        return imported
    finally:
        await db.close()


async def _import_trade_in(records: list):
    """将回收单解析结果更新到 daily_sales 表的 trade_in_qty 字段（UPSERT）"""
    if not records:
        return
    db = await aiosqlite.connect(DB_PATH_STR)
    try:
        for r in records:
            await db.execute("""
                INSERT INTO daily_sales (date, store_id, phone_sales, ncme_sales, phone_qty,
                                        key_model_qty, accessory_sales, trade_in_qty, source)
                VALUES (?, ?, 0, 0, 0, 0, 0, ?, 'eboss')
                ON CONFLICT(date, store_id) DO UPDATE SET trade_in_qty=excluded.trade_in_qty
            """, (r["date"], r["store_id"], r["trade_in_qty"]))
        await db.commit()
        logger.info(f"回收单导入: UPSERT {len(records)}条记录")
    finally:
        await db.close()


async def _import_eboss_inventory(records: list) -> int:
    """将库存汇总解析结果写入 inventory 表"""
    if not records:
        return 0
    db = await aiosqlite.connect(DB_PATH_STR)
    db.row_factory = aiosqlite.Row
    try:
        imported = 0
        for rec in records:
            store_name = rec.get("store_name", "")
            store_id = match_store_id(store_name)
            if not store_id:
                continue

            await db.execute("""
                INSERT INTO inventory (store_id, model_code, color, spec, qty, updated_at)
                VALUES (?, ?, ?, ?, ?, datetime('now', 'localtime'))
                ON CONFLICT(store_id, model_code, color, spec)
                DO UPDATE SET qty=excluded.qty, updated_at=excluded.updated_at
            """, (store_id, rec["model_code"], rec.get("color", ""),
                  rec.get("spec", ""), rec.get("qty", 0)))
            imported += 1
        await db.commit()
        return imported
    finally:
        await db.close()


async def _save_eboss_sync_log(filename, file_type, file_date, file_mtime,
                                parsed, imported, status, error_msg=""):
    """记录 eBoss 同步日志"""
    db = await aiosqlite.connect(DB_PATH_STR)
    try:
        await db.execute("""
            INSERT INTO eboss_sync_log (file_name, file_type, file_date, file_mtime,
                                       records_parsed, records_imported, status, error_msg)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (filename, file_type, file_date, file_mtime, parsed, imported, status, error_msg))
        await db.commit()
    finally:
        await db.close()


# ==================== 调度器管理 ====================

def start_scheduler():
    """启动定时任务调度器"""
    log_dir = os.path.join(PROJECT_DIR, "logs")
    os.makedirs(log_dir, exist_ok=True)

    scheduler.add_job(
        task_inventory_alert,
        CronTrigger(day_of_week="mon,thu", hour=9, minute=30),
        id="inventory_alert",
        name="库存预警检查(周一/四9:30)",
        replace_existing=True
    )

    scheduler.add_job(
        task_price_check,
        CronTrigger(hour=10, minute=0),
        id="price_check",
        name="价格检查(10:00)",
        replace_existing=True
    )

    scheduler.add_job(
        task_backup_database,
        CronTrigger(hour=3, minute=0),
        id="db_backup",
        name="数据库备份(3:00)",
        replace_existing=True
    )

    scheduler.add_job(
        task_scan_eboss_dir,
        CronTrigger(hour="*", minute=0),  # 每小时整点
        id="eboss_scan",
        name="eBoss目录扫描(每小时)",
        replace_existing=True
    )

    scheduler.start()
    logger.info("定时任务调度器已启动")
    logger.info(f"已注册任务: {[job.id for job in scheduler.get_jobs()]}")

    for job in scheduler.get_jobs():
        logger.info(f"  {job.name} → 下次执行: {job.next_run_time}")


def stop_scheduler():
    """停止调度器"""
    if scheduler.running:
        scheduler.shutdown(wait=False)
        logger.info("调度器已停止")


def get_scheduler_info() -> dict:
    """获取调度器状态信息（供API使用）"""
    jobs = []
    for job in scheduler.get_jobs():
        jobs.append({
            "id": job.id,
            "name": job.name,
            "next_run": str(job.next_run_time) if job.next_run_time else None,
        })
    return {
        "running": scheduler.running,
        "jobs": jobs
    }
