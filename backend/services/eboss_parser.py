"""
eBoss 导出文件解析服务
支持文件类型：
  - 零售汇总*.xls/xlsx  → 按门店/日期/品类汇总到 daily_sales 格式
  - 零售明细查询*.xlsx   → 带品名的订单明细（推荐，最准确）
  - 库存汇总*.xls/xlsx  → 四系列库存数据 → inventory 表
  - 零售单*.xlsx         → 零售订单明细（无品名，辅助）
"""
import os
import re
import logging
from datetime import datetime, date
from typing import List, Dict, Optional
from collections import defaultdict, Counter

from backend.logger import scheduler_logger
from backend.config import EBOSS_SCAN_DIR

logger = scheduler_logger

# eBoss 导出目录
EBOSS_DIR = EBOSS_SCAN_DIR

# 门店别名映射（eBoss 店名 → 数据库 store_id）
# 注意：eBoss 导出文件中的门店名可能与数据库不完全一致，需逐一核对
STORE_NAME_MAP = {
    "万象城三星授权旗舰店": 1,
    "华润万象汇三星授权体验店": 2,
    "兴义梦乐城三星授权体验店": 3,
    "遵义吾悦三星授权体验店": 4,
    "曲靖万达三星授权体验店": 5,
    "六盘水三星授权体验店": 6,
    "龙湾万达三星授权体验店": 7,   # eBoss 文件里是"授权体验店"不是"专卖店"
    "云南昭通三星授权体验店": 8,
    "清镇吾悦三星授权体验店": 9,
    "安顺三星授权体验店": 10,
    "蒙自店": 11,
}
# 模糊匹配关键词（eBoss 门店名中含有关键词即可匹配）
STORE_KEYWORDS = {
    1: ["万象城"],
    2: ["万象汇"],
    3: ["兴义"],
    4: ["遵义"],
    5: ["曲靖"],
    6: ["六盘水"],
    7: ["龙湾"],
    8: ["昭通"],
    9: ["清镇"],
    10: ["安顺"],
    11: ["蒙自"],
}

# 四系列重点机型编码（用于库存监控）
FOUR_SERIES_PATTERNS = ["S9420", "S9470", "S9480", "ZFOLD7", "ZFLIP7", "W26"]

# 月度销售重点机型（仅 W26）
KEY_MODEL_PATTERNS = ["W26"]


def match_store_id(store_name: str) -> Optional[int]:
    """匹配 eBoss 门店名到数据库 store_id"""
    if not store_name:
        return None
    # 精确匹配
    if store_name in STORE_NAME_MAP:
        return STORE_NAME_MAP[store_name]
    # 关键词匹配
    for sid, keywords in STORE_KEYWORDS.items():
        for kw in keywords:
            if kw in store_name:
                return sid
    return None


def is_key_model(model_str: str) -> bool:
    """判断是否为重点机型（仅 W26）
    
    正确做法：不去空格，直接在原始字符串上匹配
    W26 后面允许跟空格、括号、汉字、字母（用于容量描述），但不能直接跟数字
    例：'三星W26 512GB' → W26后跟空格，是重点机型
         'W2600'       → W26后跟0（数字），不是重点机型
    """
    if not model_str:
        return False
    upper = model_str.upper()
    # W26 后面不能直接跟数字（避免 W2600 误判）
    return bool(re.search(r'W26(?!\d)', upper))


def detect_file_type(filename: str) -> Optional[str]:
    """根据文件名检测 eBoss 报表类型"""
    if filename.startswith("库存汇总（带串号）"):
        return None  # 无颜色/规格列，暂不处理
    if filename.startswith("零售明细查询") and filename.endswith(".xlsx"):
        return "retail_detail"
    if filename.startswith("库存汇总") and filename.endswith(".xls"):
        return "inventory"
    elif filename.startswith("库存汇总") and filename.endswith(".xlsx"):
        return "inventory_xlsx"
    elif filename.startswith("零售汇总") and filename.endswith(".xls"):
        return "retail_summary"
    elif filename.startswith("零售汇总") and filename.endswith(".xlsx"):
        return "retail_summary_xlsx"
    elif filename.startswith("零售单") and filename.endswith(".xlsx"):
        return "retail_order"
    elif filename.startswith("回收单") and filename.endswith(".xlsx"):
        return "trade_in"
    elif filename.endswith(".cub"):
        return None  # 忽略 .cub 文件
    return None


# ==================== 零售汇总解析 ====================

def parse_retail_summary(file_path: str) -> List[Dict]:
    """
    解析 eBoss 零售汇总文件（.xls 或 .xlsx）
    返回按 (门店, 日期) 汇总的销售数据列表

    每条记录格式：
    {
        "store_id": int,
        "date": "YYYY-MM-DD",
        "phone_sales": float,       # 手机成交金额
        "ncme_sales": float,        # NCME 成交金额
        "phone_qty": int,           # 手机零售数量
        "key_model_qty": int,       # 重点机型数量
        "accessory_sales": float,   # 配件成交金额
        "trade_in_qty": int,        # 回收台量（含以旧换新的手机台量）
        "total_sales": float,       # 总成交金额
    }
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".xls":
            return _parse_retail_summary_xls(file_path)
        elif ext == ".xlsx":
            return _parse_retail_summary_xlsx(file_path)
    except Exception as e:
        logger.error(f"解析零售汇总失败 {file_path}: {e}", exc_info=True)
        return []
    return []


def _parse_retail_summary_xls(file_path: str) -> List[Dict]:
    """解析 .xls 格式零售汇总"""
    import xlrd
    wb = xlrd.open_workbook(file_path, formatting_info=True)
    ws = wb.sheet_by_index(0)

    # 找表头行
    header_row = 0
    for r in range(min(5, ws.nrows)):
        val = str(ws.cell_value(r, 0)).strip()
        if val == "店仓名称":
            header_row = r
            break

    # 列索引映射
    COL_STORE = 0      # 店仓名称
    COL_DATE = 1       # 日期
    COL_ATTR = 2       # 货品属性
    COL_CATEGORY = 3   # 一级(品类)
    COL_MODEL = 4      # 四级(机型)
    COL_QTY = 7        # 零售数量
    COL_DEAL_AMT = 10  # 成交金额
    COL_PAY_TYPE = 6   # 收款类型

    # 按门店+日期汇总
    summary = defaultdict(lambda: {
        "phone_sales": 0, "ncme_sales": 0, "phone_qty": 0,
        "key_model_qty": 0, "accessory_sales": 0, "trade_in_qty": 0,
        "total_sales": 0,
    })

    parsed_rows = 0
    for r in range(header_row + 1, ws.nrows):
        store_name = str(ws.cell_value(r, COL_STORE)).strip()
        date_raw = str(ws.cell_value(r, COL_DATE)).strip()
        attr = str(ws.cell_value(r, COL_ATTR)).strip() if COL_ATTR < ws.ncols else ""
        category = str(ws.cell_value(r, COL_CATEGORY)).strip() if COL_CATEGORY < ws.ncols else ""
        model = str(ws.cell_value(r, COL_MODEL)).strip() if COL_MODEL < ws.ncols else ""
        pay_type = str(ws.cell_value(r, COL_PAY_TYPE)).strip() if COL_PAY_TYPE < ws.ncols else ""

        try:
            qty = float(ws.cell_value(r, COL_QTY)) if COL_QTY < ws.ncols else 0
        except (ValueError, TypeError):
            qty = 0
        try:
            deal_amt = float(ws.cell_value(r, COL_DEAL_AMT)) if COL_DEAL_AMT < ws.ncols else 0
        except (ValueError, TypeError):
            deal_amt = 0

        if not store_name or not date_raw:
            continue

        store_id = match_store_id(store_name)
        if not store_id:
            continue

        # 解析日期格式: 20260501 → 2026-05-01
        try:
            dt = datetime.strptime(date_raw, "%Y%m%d")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

        key = (store_id, date_str)
        s = summary[key]

        # 判断品类
        is_phone = category == "移动通讯类"
        is_accessory = category == "配件类"
        is_screen = category == "碎屏业务"
        is_key = is_key_model(model)

        # NCME 判断：货品属性含"小范总库"或收款含"NCME"（仅非手机项）
        is_ncme = "NCME" in attr.upper() or "NCME" in pay_type.upper()

        # 累加
        if is_phone and not is_screen:
            s["phone_sales"] += deal_amt
            s["phone_qty"] += int(qty)
            if is_key:
                s["key_model_qty"] += int(qty)
            # 手机不再计入 NCME，避免重复
        elif is_ncme and not is_phone:
            s["ncme_sales"] += deal_amt
        elif is_accessory:
            s["accessory_sales"] += deal_amt

        s["total_sales"] += deal_amt
        parsed_rows += 1

    logger.info(f"零售汇总解析: {file_path}, 总行={ws.nrows - header_row - 1}, 有效={parsed_rows}, 门店日期组合={len(summary)}")

    # 转为列表
    results = []
    for (store_id, date_str), data in summary.items():
        results.append({
            "store_id": store_id,
            "date": date_str,
            "phone_sales": round(data["phone_sales"], 2),
            "ncme_sales": round(data["ncme_sales"], 2),
            "phone_qty": data["phone_qty"],
            "key_model_qty": data["key_model_qty"],
            "accessory_sales": round(data["accessory_sales"], 2),
            "trade_in_qty": data["trade_in_qty"],
            "total_sales": round(data["total_sales"], 2),
        })
    return results


def _parse_retail_summary_xlsx(file_path: str) -> List[Dict]:
    """解析 .xlsx 格式零售汇总"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    COL_STORE = 0
    COL_DATE = 1
    COL_ATTR = 2
    COL_CATEGORY = 3
    COL_MODEL = 4
    COL_QTY = 7
    COL_DEAL_AMT = 10
    COL_PAY_TYPE = 6

    summary = defaultdict(lambda: {
        "phone_sales": 0, "ncme_sales": 0, "phone_qty": 0,
        "key_model_qty": 0, "accessory_sales": 0, "trade_in_qty": 0,
        "total_sales": 0,
    })

    parsed_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(COL_STORE, COL_DATE, COL_CATEGORY):
            continue

        store_name = str(row[COL_STORE] or "").strip()
        date_raw = str(row[COL_DATE] or "").strip()
        attr = str(row[COL_ATTR] or "").strip() if COL_ATTR < len(row) else ""
        category = str(row[COL_CATEGORY] or "").strip()
        model = str(row[COL_MODEL] or "").strip() if COL_MODEL < len(row) else ""
        pay_type = str(row[COL_PAY_TYPE] or "").strip() if COL_PAY_TYPE < len(row) else ""

        try:
            qty = float(row[COL_QTY]) if COL_QTY < len(row) and row[COL_QTY] else 0
        except (ValueError, TypeError):
            qty = 0
        try:
            deal_amt = float(row[COL_DEAL_AMT]) if COL_DEAL_AMT < len(row) and row[COL_DEAL_AMT] else 0
        except (ValueError, TypeError):
            deal_amt = 0

        if not store_name or not date_raw:
            continue

        store_id = match_store_id(store_name)
        if not store_id:
            continue

        try:
            dt = datetime.strptime(date_raw, "%Y%m%d")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

        key = (store_id, date_str)
        s = summary[key]

        is_phone = category == "移动通讯类"
        is_accessory = category == "配件类"
        is_key = is_key_model(model)
        is_ncme = "NCME" in attr.upper() or "NCME" in pay_type.upper()

        if is_phone:
            s["phone_sales"] += deal_amt
            s["phone_qty"] += int(qty)
            if is_key:
                s["key_model_qty"] += int(qty)
            # 手机不再计入 NCME
        elif is_ncme and not is_phone:
            s["ncme_sales"] += deal_amt
        elif is_accessory:
            s["accessory_sales"] += deal_amt

        s["total_sales"] += deal_amt
        parsed_rows += 1

    logger.info(f"零售汇总(xlsx)解析: {file_path}, 有效={parsed_rows}, 组合={len(summary)}")

    results = []
    for (store_id, date_str), data in summary.items():
        results.append({
            "store_id": store_id,
            "date": date_str,
            "phone_sales": round(data["phone_sales"], 2),
            "ncme_sales": round(data["ncme_sales"], 2),
            "phone_qty": data["phone_qty"],
            "key_model_qty": data["key_model_qty"],
            "accessory_sales": round(data["accessory_sales"], 2),
            "trade_in_qty": data["trade_in_qty"],
            "total_sales": round(data["total_sales"], 2),
        })
    return results


# ==================== 零售单解析 ====================

def parse_retail_order(file_path: str) -> List[Dict]:
    """
    解析 eBoss 零售单文件（.xlsx）
    返回订单明细列表（每条为一个零售单）

    字段映射：
    - 单据日期 → date
    - 零售店铺 → store_id
    - 零售数量 → qty
    - 零售金额 → amount
    - 成交金额 → deal_amount
    - 收款方式 → pay_method
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    COL_ORDER_NO = 1   # 单据编号
    COL_DATE = 2        # 单据日期
    COL_STORE = 3       # 零售店铺
    COL_QTY = 8         # 零售数量
    COL_AMOUNT = 9      # 零售金额
    COL_DEAL = 10       # 成交金额
    COL_PAY = 23        # 收款方式 (X列, index=23)
    COL_VOID = 33       # 作否 (AH列, index=33)

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= COL_VOID:
            continue

        store_name = str(row[COL_STORE] or "").strip()
        date_raw = str(row[COL_DATE] or "").strip()
        is_void = str(row[COL_VOID] or "").strip()

        if not store_name or not date_raw or is_void == "是":
            continue

        store_id = match_store_id(store_name)
        if not store_id:
            continue

        try:
            dt = datetime.strptime(date_raw, "%Y%m%d")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

        try:
            qty = int(float(row[COL_QTY] or 0))
        except (ValueError, TypeError):
            qty = 0
        try:
            amount = float(row[COL_AMOUNT] or 0)
        except (ValueError, TypeError):
            amount = 0
        try:
            deal = float(row[COL_DEAL] or 0)
        except (ValueError, TypeError):
            deal = 0
        pay = str(row[COL_PAY] or "").strip() if len(row) > COL_PAY else ""

        orders.append({
            "store_id": store_id,
            "date": date_str,
            "order_no": str(row[COL_ORDER_NO] or ""),
            "qty": qty,
            "amount": round(amount, 2),
            "deal_amount": round(deal, 2),
            "pay_method": pay,
        })

    logger.info(f"零售单解析: {file_path}, 订单数={len(orders)}")
    return orders


# ==================== 零售明细查询解析（推荐格式）====================

# 零售明细查询 xlsx 列索引（0-based）
_RD_COL_STORE = 1        # 店铺
_RD_COL_WAREHOUSE = 2     # 专卖店仓（小范总库=NCME）
_RD_COL_DATE = 3          # 单据日期 (YYYYMMDD)
_RD_COL_STATUS = 8        # 状态（正常=有效）
_RD_COL_PRODUCT = 10      # 品名（含机型编码）
_RD_COL_SERIAL = 12       # 带串号（是=实体机）
_RD_COL_QTY = 13          # 零售数量
_RD_COL_RETAIL_AMT = 17   # 零售金额
_RD_COL_DEAL_AMT = 18     # 成交金额
_RD_COL_REMARK = 28       # 备注（含回收/换新信息）

# NCME 产品识别正则（手表、耳机、平板、手环、戒指）
# 注意：WATCH\b 不能匹配 Watch7（数字不是 word boundary），改用 WATCH(?:\s|\d|$|_)
_NCME_PRODUCT_RE = re.compile(
    r'(手表|WATCH(?:\s|\d|$|_|-)|'  # Galaxy Watch / Watch7 / Watch FE
    r'手环|FIT\d|RING(?:\s|\d|$)|'  # Galaxy Fit / Ring
    r'耳机|BUDS|'                    # Galaxy Buds
    r'平板|SM-X\d)',                 # Galaxy Tab (SM-X系列)
    re.IGNORECASE
)

# 手机型号识别正则（匹配品名中的三星机型编码，不含 SM-W/SM-X 等非手机系列）
_PHONE_MODEL_RE = re.compile(
    r'(S9420|S9470|S9480|ZFOLD7|ZFLIP7|W26|'
    r'S9460|S9370|S9380|S9360|S9350|S9340|'
    r'F9660|F7660|F9680|M-F\d+|'
    r'A\d{4}\s*[\(（]\s*\d)'
)

# 所有三星手机型号编码（非配件，不含 SM-X 平板系列）
_ALL_PHONE_MODEL_RE = re.compile(
    r'(?:三星|Samsung).*?(?:S\d{4}|ZFOLD\d+|ZFLIP\d+|W\d+|A\d{4}|SM-9\d|SM-7\d)\s*[\(（]'
)


def parse_retail_detail(file_path: str) -> List[Dict]:
    """
    解析 eBoss 零售明细查询文件（带品名的订单明细，推荐格式）
    返回按 (门店, 日期) 汇总的销售数据列表

    分类规则（按优先级）：
    1. NCME：品名匹配手表/耳机/平板/手环/戒指 → ncme_sales
    2. 手机：带串号=是 OR 品名匹配手机型号编码 → phone_sales
    3. 重点机型（仅 W26）：品名含 W26 → key_model_qty
    4. 其他：→ accessory_sales
    - 跳过：售后仓、订金、碎屏服务、成交金额=0
    - 回收台量：不从本表获取，统一为 0（由独立数据源提供）
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    summary = defaultdict(lambda: {
        "phone_sales": 0, "ncme_sales": 0, "phone_qty": 0,
        "key_model_qty": 0, "accessory_sales": 0, "trade_in_qty": 0,
        "total_sales": 0,
    })

    parsed_rows = 0
    skipped_reasons = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(_RD_COL_STORE, _RD_COL_DATE, _RD_COL_STATUS,
                          _RD_COL_PRODUCT, _RD_COL_DEAL_AMT, _RD_COL_REMARK):
            continue

        store_name = str(row[_RD_COL_STORE] or "").strip()
        warehouse = str(row[_RD_COL_WAREHOUSE] or "").strip()
        date_raw = str(row[_RD_COL_DATE] or "").strip()
        status = str(row[_RD_COL_STATUS] or "").strip()
        product = str(row[_RD_COL_PRODUCT] or "").strip()
        has_serial = str(row[_RD_COL_SERIAL] or "").strip()
        remark = str(row[_RD_COL_REMARK] or "").strip()

        # 过滤无效行
        if not store_name or not date_raw:
            skipped_reasons["无店铺/日期"] += 1
            continue
        # 退货行保留参与计算（含负数金额/数量，自然抵消原销售额）
        # 不能跳过，否则"卖了又退"的商品仍被计入销售额

        # 排除售后仓
        if "售后" in store_name:
            skipped_reasons["售后仓"] += 1
            continue

        # 排除订金
        if "订金" in product:
            skipped_reasons["订金"] += 1
            continue

        # 解析成交金额
        try:
            deal_amt = float(row[_RD_COL_DEAL_AMT] or 0)
        except (ValueError, TypeError):
            deal_amt = 0

        # 跳过零金额行
        if deal_amt == 0:
            skipped_reasons["金额=0"] += 1
            continue

        # 解析数量
        try:
            qty = int(float(row[_RD_COL_QTY] or 0))
        except (ValueError, TypeError):
            qty = 0

        # 匹配门店
        store_id = match_store_id(store_name)
        if not store_id:
            skipped_reasons[f"门店未匹配:{store_name[:10]}"] += 1
            continue

        # 解析日期
        try:
            dt = datetime.strptime(date_raw, "%Y%m%d")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            skipped_reasons[f"日期格式错:{date_raw}"] += 1
            continue

        key = (store_id, date_str)
        s = summary[key]

        # ---- 品类分类 ----
        # 核心规则：
        # 1. NCME渠道（小范总库）的非手机产品 → ncme_sales
        # 2. 平板(TAB S/SM-X)不是手机，从小范总库出 → NCME，从常规仓库出 → NCME
        # 3. 手机 = 实际手机型号（S26/W26/ZFOLD等），不含平板
        # 4. 配件 = 膜/壳/充电器等（非NCME渠道的配件）
        # 注意：手表归入NCME（从小范总库）或配件（从常规仓库），2026-06-10 修正

        # NCME渠道判断：专卖店仓 = 小范总库
        is_ncme_warehouse = (warehouse == "小范总库")

        # 平板识别（TAB S系列，SM-X编码）
        is_tablet = bool(re.search(r'TAB\s*S|SM-X\d', product, re.IGNORECASE))

        # 1. 配件：膜/壳/充电器/手表/手环等（非NCME渠道的才归配件）
        _ACCESSORY_KEYWORDS = ["膜", "壳", "充电器", "充电线", "数据线", "电池",
            "保护", "壳膜", "支架", "自拍杆", "移动电源", "充电宝",
            "电源适配", "车充", "座充", "无线充",
            "手表", "WATCH", "Watch", "Galaxy Watch", "手环"]
        is_accessory_product = any(kw in product for kw in _ACCESSORY_KEYWORDS)

        # 碎屏服务单独处理（不算任何品类）
        if "碎屏" in product:
            skipped_reasons["碎屏服务"] += 1
            continue

        # 2. 手机判断：实际手机型号 OR 带串号=是（但排除平板、手表等非手机）
        is_phone_model = bool(_ALL_PHONE_MODEL_RE.search(product.upper()))
        is_serial = (has_serial == "是")
        # 手机 = 有串号 或 匹配手机型号，但排除平板
        is_phone = (is_serial or is_phone_model) and not is_tablet and not is_accessory_product

        # 3. NCME判断：小范总库的非手机产品（手表/平板等）
        #    OR 平板（无论哪个仓库，平板都是NCME产品）
        is_ncme = (is_ncme_warehouse and not is_phone) or is_tablet

        # 4. 重点机型（仅 W26，用词边界确保 W260/W2600 不误判）
        is_key_model = bool(re.search(r'W26(?!\d)', product.upper())) and "W25" not in product.upper()

        # 分类优先级：NCME > 配件 > 手机 > 其他
        if is_ncme:
            # NCME渠道的非手机产品（手表/平板等）
            s["ncme_sales"] += deal_amt
        elif is_accessory_product and not is_ncme_warehouse:
            # 常规渠道的配件（膜/壳/充电器/手表等）
            s["accessory_sales"] += deal_amt
        elif is_phone:
            # 手机
            s["phone_sales"] += deal_amt
            s["phone_qty"] += qty
            if is_key_model:
                s["key_model_qty"] += qty
        else:
            # 其他未分类项 → accessory_sales
            s["accessory_sales"] += deal_amt

        s["total_sales"] += deal_amt
        parsed_rows += 1

    logger.info(
        f"零售明细解析: {file_path}, "
        f"数据行={ws.max_row - 1}, 有效={parsed_rows}, 组合={len(summary)}"
    )
    if skipped_reasons:
        logger.info(f"跳过明细: {dict(skipped_reasons)}")

    results = []
    for (store_id, date_str), data in summary.items():
        results.append({
            "store_id": store_id,
            "date": date_str,
            "phone_sales": round(data["phone_sales"], 2),
            "ncme_sales": round(data["ncme_sales"], 2),
            "phone_qty": data["phone_qty"],
            "key_model_qty": data["key_model_qty"],
            "accessory_sales": round(data["accessory_sales"], 2),
            "trade_in_qty": data["trade_in_qty"],
            "total_sales": round(data["total_sales"], 2),
        })
    return results


# ==================== 库存汇总解析（复用 inventory.py）====================

def parse_inventory_xls(file_path: str) -> List[Dict]:
    """
    解析 eBoss 库存汇总文件
    复用 inventory.py 的 parse_eboss_xls 逻辑
    """
    try:
        from backend.api.inventory import parse_eboss_xls
        return parse_eboss_xls(file_path)
    except Exception as e:
        logger.error(f"解析库存汇总失败 {file_path}: {e}", exc_info=True)
        return []


def parse_inventory_xlsx(file_path: str) -> List[Dict]:
    """解析 .xlsx 格式库存汇总"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 遍历所有 sheet，找包含门店数据的 sheet
    for ws in wb.worksheets:
        has_store = False
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and any(kw in cell for kw in ["万象城", "兴义", "龙湾", "六盘水"]):
                    has_store = True
                    break
            if has_store:
                break
        if has_store:
            return _parse_inventory_xlsx_sheet(ws)

    return []


def _parse_inventory_xlsx_sheet(ws) -> List[Dict]:
    """解析单个 xlsx sheet 的库存数据"""
    records = []
    cur_store, cur_model, cur_color, cur_series = "", "", "", ""

    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = [str(v or "").strip() if v is not None else "" for v in row]
        if len(vals) < 6:
            continue

        c0 = vals[0]   # 门店
        c2 = vals[2]   # 机型
        c3 = vals[3]   # 规格
        c4 = vals[4]   # 颜色
        c5 = vals[5]   # 库存量

        if c0 and c0 != "合计" and len(c0) > 1:
            cur_store = c0
        if c2 and c2 != "合计":
            cur_model = c2
            cur_series = get_series_from_model_xlsx(c2)
            cur_color = ""
        if c4 and c4 != "合计":
            cur_color = c4
        if c2 == "合计" or c0 == "合计":
            continue
        if not cur_series:
            continue

        if c3 and c3 != "合计":
            try:
                qty = int(float(c5)) if c5 else 0
            except (ValueError, TypeError):
                qty = 0
            if qty > 0:
                records.append({
                    "store_name": cur_store,
                    "model_code": normalize_model_code(cur_model),
                    "series": cur_series,
                    "color": cur_color,
                    "spec": c3,
                    "qty": qty,
                })

    logger.info(f"库存汇总(xlsx)解析: records={len(records)}")
    return records


def get_series_from_model_xlsx(model_str: str) -> str:
    """判断机型系列"""
    FOUR_SERIES = {
        'S9420': 'S26', 'S9470': 'S26', 'S9480': 'S26',
        'ZFOLD7': 'FOLD7', 'ZFLIP7': 'FLIP7', 'W26': 'W26',
    }
    for key, series in FOUR_SERIES.items():
        if key in model_str.upper().replace(" ", ""):
            return series
    return ""


def normalize_model_code(model_str: str) -> str:
    """标准化机型编码"""
    m = re.search(r"(S\d{4}|ZFOLD\d+|ZFLIP\d+|W\d+)", model_str.upper().replace(" ", ""))
    if m:
        return m.group(1)
    for key in ["S9420", "S9470", "S9480", "ZFOLD7", "ZFLIP7", "W26"]:
        if key in model_str.upper().replace(" ", ""):
            return key
    return model_str


# ==================== 目录扫描 ====================

def scan_eboss_directory(target_dir: str = EBOSS_DIR) -> List[Dict]:
    """
    扫描 eBoss 导出目录，返回可处理的文件列表
    每个文件: {filename, filetype, mtime, filepath}
    """
    if not os.path.isdir(target_dir):
        logger.warning(f"eBoss 目录不存在: {target_dir}")
        return []

    files = []
    for fname in os.listdir(target_dir):
        ftype = detect_file_type(fname)
        if not ftype or ftype == "cub":
            continue

        fpath = os.path.join(target_dir, fname)
        if not os.path.isfile(fpath):
            continue

        mtime = os.path.getmtime(fpath)
        files.append({
            "filename": fname,
            "filetype": ftype,
            "mtime": datetime.fromtimestamp(mtime).isoformat(),
            "filepath": fpath,
        })

    # 按修改时间倒序
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return files


def parse_trade_in(filepath: str) -> List[Dict]:
    """
    解析 eBoss 回收单文件（回收单*.xlsx）

    列结构：
      Col3: 单据日期 (格式 YYYYMMDD)
      Col4: 回收门店
      Col9: 回收数量
      Col18: 状态 (已付款/审核)

    返回按 (门店, 日期) 汇总的回收数据列表，每条格式：
      {"date": "YYYY-MM-DD", "store_id": int, "trade_in_qty": int}
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    grouped = defaultdict(lambda: {"trade_in_qty": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 18:
            continue

        date_raw = row[2]  # Col3: 单据日期（可能是 datetime 对象或字符串）
        store_name = str(row[3] or "").strip()  # Col4: 回收门店
        qty_raw = row[8]  # Col9: 回收数量

        if not date_raw or not qty_raw:
            continue

        try:
            qty = int(qty_raw)
        except (ValueError, TypeError):
            continue

        store_id = match_store_id(store_name)
        if store_id is None:
            logger.warning("回收单门店未匹配: %s", store_name)
            continue

        # 日期格式转换：支持 datetime 对象和 YYYYMMDD 字符串
        date_str = None
        if isinstance(date_raw, datetime):
            date_str = date_raw.strftime("%Y-%m-%d")
        elif isinstance(date_raw, str):
            date_clean = date_raw.strip()
            # YYYYMMDD 格式
            if len(date_clean) == 8 and date_clean.isdigit():
                try:
                    dt = datetime.strptime(date_clean, "%Y%m%d")
                    date_str = dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
            # YYYY-MM-DD 格式
            elif "-" in date_clean:
                try:
                    dt = datetime.strptime(date_clean[:10], "%Y-%m-%d")
                    date_str = dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
        if not date_str:
            continue

        key = (date_str, store_id)
        grouped[key]["trade_in_qty"] += qty

    records = []
    for (d, sid), vals in grouped.items():
        records.append({
            "date": d,
            "store_id": sid,
            "trade_in_qty": vals["trade_in_qty"],
        })

    records.sort(key=lambda x: (x["date"], x["store_id"]))
    logger.info("回收单解析完成: %d条记录, %d个门店×日期组合", len(records), len(records))
    return records
